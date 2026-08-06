# -*- coding: utf-8 -*-
"""
test_regression.py —— 回归测试

模拟数据的表头、列顺序、正负号约定全部按五份真实账单构造。
每次改完 customer_config.py 或核心逻辑，跑一次 `python test_regression.py`，
确保修好一个客户不会弄坏另一个。

覆盖场景
--------
  1. 补充文件列数/顺序与主文件不同（按表头映射，不错位）
  2. 补充订单覆盖主文件；补充文件新增订单（漏单）
  3. 完全一致的重复行被丢弃
  4. 七月创建、八月完成的快递计入七月
  5. 六月创建、七月完成的快递不计入七月
  6. 卡派自提按【完成时间】归属账期
  7. 订单处理费为 0 但有物流费 —— 必须保留
  8. 无燃油客户（GUXE）无需费率直接运行
  9. 燃油边界日期（费率周首尾日）
 10. 复合订单多个 SKU 全部保留
 11. PlatformFee 纳入物流费用
 12. 手工计费列纳入费用总计
 13. 1970 年日期阻断计算
 14. Decimal 逐单 ROUND_HALF_UP
 15. 【汇总】模板按标签写入，前八行与第二段不被破坏
 16. 账单原值 vs 明细合计 对账（漏单检出）
 17. 燃油费率区间校验（重叠 / 空档 / 未覆盖）
 18. 燃油费率记忆：保存、读取、缺周识别
 19. 从源账单反推燃油费率
"""

import os
import shutil
import sys
import tempfile
from datetime import date, datetime
from decimal import Decimal

import pandas as pd
from openpyxl import load_workbook

import billing_core as core
import customer_config as cfgmod
import fuel_rates as fr
import rate_store
import sheet_merge as sm
from money import round_money, to_decimal

PASSED, FAILED = [], []


def check(name, condition, detail=""):
    if condition:
        PASSED.append(name)
        print(f"  [PASS] {name}")
    else:
        FAILED.append((name, detail))
        print(f"  [FAIL] {name} {detail}")


# --------------------------------------------------------------------------
# 构造测试数据（列名/顺序/正负号与真实账单一致）
# --------------------------------------------------------------------------

MAIN_COLUMNS = [
    "客户编码", "出库仓库", "订单编号", "外部订单号", "订单类型", "尾程物流",
    "创建时间", "发货时间", "完成时间", "SKU", "订单备注", "币种",
    "订单处理费", "自有运单计费", "复合订单处理费", "物流费用",
    "贴标/换标(手动计费)", "拣货费(手动计费)", "自提出库费(手动计费)", "出库打托费(手动计费)",
    "ServiceWay", "Price", "PlatformFee", "Zone", "计费重",
    "住宅地址附加费", "偏远附加费", "基础运费", "操作附加费-尺寸",
    "燃油附加费", "操作附加费", "超尺寸附加费", "费用总计",
]

# 汇总模板（与真实账单一致：第 8 行是合计、9~16 行是费用、19 行是期末余额）
SUMMARY_LABELS = [
    "客户编号", "客户昵称", "公司名称", "账单币种", "账单起止日期",
    "账单初账户余额", "减 本期账单总支出",
    "出库订单费", "入库订单费", "工单处理费", "退件处理费",
    "仓租费", "询价打单费", "核账补收费", "其他费",
    "加 账单充值总额", "退款/返款", "账期末账号余额",
]


def order_row(order_id, sku, last_mile, created, finished, order_type="单件代发",
              proc=-2.2, own=0, comp=0, label=0, pick=0, self_out=0, pallet=0,
              base=10.0, resi=2.5, remote=0, oper=0, oper_size=0, oversize=0,
              src_fuel=0, platform=0, remark="", total=None, logistics=None):
    freight = base + resi + remote + oper + oper_size + oversize
    if logistics is None:
        logistics = -round(freight + src_fuel + platform, 2) if freight else 0
    if total is None:
        total = round(proc + own + comp + label + pick + self_out + pallet + logistics, 2)
    return {
        "客户编码": "TEST", "出库仓库": "USGA01", "订单编号": order_id,
        "外部订单号": f"E{order_id}", "订单类型": order_type, "尾程物流": last_mile,
        "创建时间": created, "发货时间": created, "完成时间": finished,
        "SKU": sku, "订单备注": remark, "币种": "USD",
        "订单处理费": proc, "自有运单计费": own, "复合订单处理费": comp, "物流费用": logistics,
        "贴标/换标(手动计费)": label, "拣货费(手动计费)": pick,
        "自提出库费(手动计费)": self_out, "出库打托费(手动计费)": pallet,
        "ServiceWay": "fedex_ground", "Price": freight + src_fuel, "PlatformFee": platform,
        "Zone": "4", "计费重": "20lbs",
        "住宅地址附加费": resi, "偏远附加费": remote, "基础运费": base,
        "操作附加费-尺寸": oper_size, "燃油附加费": src_fuel,
        "操作附加费": oper, "超尺寸附加费": oversize, "费用总计": total,
    }


def summary_frame(code, outbound_total):
    values = {
        "客户编号": code, "客户昵称": "test", "公司名称": "****",
        "账单币种": "USD", "账单起止日期": "2026-07-01-2026-07-31",
        "账单初账户余额": -1000.0, "减 本期账单总支出": 0,
        "出库订单费": outbound_total, "入库订单费": 0, "工单处理费": 0, "退件处理费": 0,
        "仓租费": -100.55, "询价打单费": 0, "核账补收费": 0, "其他费": 0,
        "加 账单充值总额": 0, "退款/返款": 0, "账期末账号余额": 0,
    }
    rows = [{"类别": lbl, "金额": values[lbl]} for lbl in SUMMARY_LABELS]
    rows.append({"类别": None, "金额": None})              # 第 20 行空行
    rows.append({"类别": "客户编号", "金额": code})         # 第二段：物流账户
    rows.append({"类别": "账期末账号余额", "金额": 0})
    return pd.DataFrame(rows)


def build_main_file(path, outbound_total_in_summary=None):
    rows = [
        order_row("A001", "SKU1", "平台快递", datetime(2026, 7, 6, 10, 0), datetime(2026, 7, 8)),
        order_row("A002", "SKU1", "平台快递", datetime(2026, 7, 30), datetime(2026, 8, 2)),
        order_row("A003", "SKU1", "平台快递", datetime(2026, 6, 28), datetime(2026, 7, 1)),
        order_row("A004", "SKU1", "卡派自提", datetime(2026, 6, 30), datetime(2026, 7, 3),
                  base=0, resi=0),
        order_row("A005", "SKU1", "平台快递", datetime(2026, 7, 7), datetime(2026, 7, 9), proc=0),
        # 7/12 是费率周（6/29~7/5、7/6~7/12）的末日
        order_row("A006", "SKU1", "平台快递", datetime(2026, 7, 12, 23, 30), datetime(2026, 7, 14)),
        order_row("A007", "SKU-A", "平台快递", datetime(2026, 7, 15), datetime(2026, 7, 16),
                  order_type="复合代发", comp=-2.0),
        order_row("A007", "SKU-B", "平台快递", datetime(2026, 7, 15), datetime(2026, 7, 16),
                  order_type="复合代发", proc=0, base=0, resi=0),
        order_row("A008", "SKU1", "平台快递", datetime(2026, 7, 20), datetime(2026, 7, 21),
                  platform=3.33, label=-1.11, pick=-0.55),
        order_row("A009", "SKU1", "平台快递", datetime(2026, 7, 21), datetime(2026, 7, 22),
                  remark="测试订单"),
        order_row("A010", "SKU1", "平台快递", datetime(2026, 7, 22), datetime(2026, 7, 23),
                  base=99, resi=0),
        order_row("A001", "SKU1", "平台快递", datetime(2026, 7, 6, 10, 0), datetime(2026, 7, 8)),
    ]
    df_out = pd.DataFrame(rows, columns=MAIN_COLUMNS)
    df_rent = pd.DataFrame([
        {"时间": datetime(2026, 7, 1), "仓库": "W1", "实际费用": -100.55},
        {"时间": datetime(2026, 6, 1), "仓库": "W1", "实际费用": -999.99},
    ])
    df_inbound = pd.DataFrame([{"入库编号": "I1", "费用": 0}])
    total = outbound_total_in_summary
    if total is None:
        total = round(sum(r["费用总计"] for r in rows), 2)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary_frame("TEST", total).to_excel(writer, sheet_name="汇总", index=False)
        df_rent.to_excel(writer, sheet_name="仓租", index=False)
        df_inbound.to_excel(writer, sheet_name="入库详情", index=False)
        df_out.to_excel(writer, sheet_name="出库订单", index=False)


def build_supplement_file(path):
    """补充文件：列更少、顺序不同（模拟 50 列 vs 54 列）。"""
    cols = ["尾程物流", "订单编号", "SKU", "订单类型", "创建时间", "完成时间",
            "订单处理费", "基础运费", "住宅地址附加费", "PlatformFee", "物流费用", "费用总计"]
    rows = [
        {"尾程物流": "平台快递", "订单编号": "A010", "SKU": "SKU1", "订单类型": "单件代发",
         "创建时间": datetime(2026, 7, 22), "完成时间": datetime(2026, 7, 23),
         "订单处理费": -2.2, "基础运费": 12, "住宅地址附加费": 0,
         "PlatformFee": 0, "物流费用": -15.0, "费用总计": -17.2},
        {"尾程物流": "平台快递", "订单编号": "A011", "SKU": "SKU1", "订单类型": "单件代发",
         "创建时间": datetime(2026, 7, 25), "完成时间": datetime(2026, 7, 26),
         "订单处理费": -2.2, "基础运费": 20, "住宅地址附加费": 0,
         "PlatformFee": 0, "物流费用": -25.0, "费用总计": -27.2},
    ]
    pd.DataFrame(rows, columns=cols).to_excel(path, sheet_name="出库订单", index=False)


def build_guxe_file(path):
    """无燃油客户：全部自有快递，且完全没有运费明细列。"""
    cols = ["客户编码", "订单编号", "订单类型", "尾程物流", "创建时间", "完成时间", "SKU",
            "订单处理费", "自有运单计费", "复合订单处理费", "物流费用", "费用总计"]
    rows = [
        {"客户编码": "GUXE", "订单编号": "G001", "订单类型": "单件代发", "尾程物流": "自有快递",
         "创建时间": datetime(2026, 7, 3), "完成时间": datetime(2026, 7, 4), "SKU": "SKU1",
         "订单处理费": -0.8, "自有运单计费": -1.5, "复合订单处理费": 0,
         "物流费用": 0, "费用总计": -2.3},
        {"客户编码": "GUXE", "订单编号": "G002", "订单类型": "复合代发", "尾程物流": "自有快递",
         "创建时间": datetime(2026, 7, 10), "完成时间": datetime(2026, 7, 11), "SKU": "SKU-A",
         "订单处理费": -0.8, "自有运单计费": -1.5, "复合订单处理费": 0,
         "物流费用": 0, "费用总计": -2.3},
        {"客户编码": "GUXE", "订单编号": "G002", "订单类型": "复合代发", "尾程物流": "自有快递",
         "创建时间": datetime(2026, 7, 10), "完成时间": datetime(2026, 7, 11), "SKU": "SKU-B",
         "订单处理费": 0, "自有运单计费": 0, "复合订单处理费": 0,
         "物流费用": 0, "费用总计": 0},
    ]
    df_rent = pd.DataFrame([{"时间": datetime(2026, 7, 1), "仓库": "W1", "实际费用": -50.0}])
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary_frame("GUXE", -4.6).to_excel(writer, sheet_name="汇总", index=False)
        df_rent.to_excel(writer, sheet_name="仓租", index=False)
        pd.DataFrame(rows, columns=cols).to_excel(writer, sheet_name="出库订单", index=False)


PERIODS_JULY = [
    fr.FuelPeriod(date(2026, 6, 29), date(2026, 7, 5), Decimal("0.2525")),
    fr.FuelPeriod(date(2026, 7, 6), date(2026, 7, 12), Decimal("0.2500")),
    fr.FuelPeriod(date(2026, 7, 13), date(2026, 7, 19), Decimal("0.2500")),
    fr.FuelPeriod(date(2026, 7, 20), date(2026, 7, 26), Decimal("0.2525")),
    fr.FuelPeriod(date(2026, 7, 27), date(2026, 8, 2), Decimal("0.2575")),
]


def read_sheet(path, name):
    return pd.read_excel(path, sheet_name=name)


# --------------------------------------------------------------------------
# 测试
# --------------------------------------------------------------------------

def test_main(tmpdir):
    print("\n--- 场景一：MPAV 规则（有燃油、主文件 + 补充文件）---")
    main_path = os.path.join(tmpdir, "main.xlsx")
    supp_path = os.path.join(tmpdir, "supp_week4.xlsx")
    out_path = os.path.join(tmpdir, "main_已处理.xlsx")

    build_main_file(main_path)
    build_supplement_file(supp_path)

    result = core.run_pipeline(
        main_path=main_path, supplement_paths=[supp_path],
        year=2026, month=7, output_path=out_path,
        customer_code="MPAV", fuel_periods=PERIODS_JULY, log=lambda _t: None,
    )

    check("输出文件已生成", os.path.exists(out_path))
    oc = read_sheet(out_path, "逐单核对")
    ids = list(oc["订单编号"].astype(str))

    check("1. 列数/顺序不同的补充文件不会错位", "A011" in ids)
    check("2. 补充订单覆盖主文件", "A010" in ids)
    a010 = oc[oc["订单编号"] == "A010"].iloc[0]
    check("2b. 覆盖后使用补充文件的金额（燃油基数=12）",
          abs(float(a010["燃油基数"]) - 12.0) < 1e-9, f"实际 {a010['燃油基数']}")
    check("2c. 覆盖记录已留档", len(read_sheet(out_path, "补充文件覆盖记录")) >= 1)
    check("3. 完全一致的重复行被丢弃", len(read_sheet(out_path, "重复订单")) >= 1)

    check("4. 七月创建、八月完成的快递计入七月", "A002" in ids)
    check("5. 六月创建、七月完成的快递不计入七月", "A003" not in ids)
    cross = read_sheet(out_path, "跨月订单")
    check("5b. 跨月订单留档", "A003" in list(cross.get("订单编号", pd.Series(dtype=str)).astype(str)))

    check("6. 卡派自提按完成时间归属七月", "A004" in ids)
    a004 = oc[oc["订单编号"] == "A004"].iloc[0]
    check("6b. 卡派自提使用【完成时间】字段", a004["账期字段"] == "完成时间", f"实际 {a004['账期字段']}")
    check("6c. 卡派自提不计燃油", abs(float(a004["实际燃油费"])) < 1e-9)

    check("7. 订单处理费为0的订单未被删除", "A005" in ids)
    a005 = oc[oc["订单编号"] == "A005"].iloc[0]
    check("7b. 处理费为0但物流费仍计算", float(a005["物流费用"]) < 0, f"实际 {a005['物流费用']}")

    a006 = oc[oc["订单编号"] == "A006"].iloc[0]
    check("9. 边界日期 7/12（周日）命中 25.00% 区间",
          a006["燃油费率"] == "25.00%", f"实际 {a006['燃油费率']}")

    check("10. 复合订单两行 SKU 都保留", ids.count("A007") == 2)
    a007 = oc[oc["订单编号"] == "A007"]
    check("10b. 只有一行标记为订单首行", list(a007["是否订单首行"]).count("是") == 1)

    a008 = oc[oc["订单编号"] == "A008"].iloc[0]
    base = to_decimal(a008["燃油基数"])
    fuel = to_decimal(a008["实际燃油费"])
    platform = to_decimal(a008["PlatformFee"])
    expect_logi = -round_money(base + fuel + platform)
    check("11. PlatformFee 已纳入物流费用",
          to_decimal(a008["物流费用"]) == expect_logi,
          f"实际 {a008['物流费用']} 期望 {expect_logi}")
    check("11b. PlatformFee 读到非零值", platform == Decimal("3.33"), f"实际 {platform}")

    expect_total = round_money(to_decimal("-2.2") + to_decimal("-1.11") + to_decimal("-0.55") + expect_logi)
    check("12. 手工计费列计入费用总计",
          to_decimal(a008["重算费用总计"]) == expect_total,
          f"实际 {a008['重算费用总计']} 期望 {expect_total}")

    check("13. 默认配置下不因备注含「测试」而误删订单", "A009" in ids,
          "默认 test_order_keywords 为空，必须保留")

    a001 = oc[oc["订单编号"] == "A001"].iloc[0]
    expected_fuel = round_money(to_decimal("12.5") * Decimal("0.25"))   # 3.125 -> 3.13
    check("14. 燃油费按 Decimal ROUND_HALF_UP 逐单取整",
          to_decimal(a001["实际燃油费"]) == expected_fuel,
          f"实际 {a001['实际燃油费']} 期望 {expected_fuel}")
    check("14b. 与银行家舍入结果不同（3.13 而非 3.12）", expected_fuel == Decimal("3.13"))

    total_recalc = sum(to_decimal(v) for v in oc["重算费用总计"])
    check("逐单合计 = 统计口径",
          round_money(total_recalc) == to_decimal(result["stats"]["重算费用总计"]),
          f"{total_recalc} vs {result['stats']['重算费用总计']}")

    wb = load_workbook(out_path)
    check("输出包含全部核账工作表",
          all(s in wb.sheetnames for s in core.REPORT_SHEETS),
          f"缺少 {[s for s in core.REPORT_SHEETS if s not in wb.sheetnames]}")
    check("燃油费率表已写入", core.SHEET_FUEL_RATE in wb.sheetnames)
    return out_path


def test_test_order_rule(tmpdir):
    """默认不配测试单关键词（防误删）；一旦配置，测试单必须被剔除并留档。"""
    print("\n--- 场景一之二：测试订单剔除规则 ---")
    main_path = os.path.join(tmpdir, "testorder.xlsx")
    out_path = os.path.join(tmpdir, "testorder_已处理.xlsx")
    build_main_file(main_path)

    original = cfgmod.CUSTOMER_OVERRIDES["MPAV"].get("drop_rules")
    cfgmod.CUSTOMER_OVERRIDES["MPAV"]["drop_rules"] = {"test_order_keywords": ["测试"]}
    try:
        core.run_pipeline(main_path=main_path, year=2026, month=7, output_path=out_path,
                          customer_code="MPAV", fuel_periods=PERIODS_JULY, log=lambda _t: None)
        oc = read_sheet(out_path, "逐单核对")
        manual = read_sheet(out_path, "手工调整")
        check("13d. 配置关键词后测试订单被剔除",
              "A009" not in list(oc["订单编号"].astype(str)))
        check("13e. 测试订单留档到「手工调整」",
              "A009" in list(manual.get("订单编号", pd.Series(dtype=str)).astype(str)))
    finally:
        if original is None:
            cfgmod.CUSTOMER_OVERRIDES["MPAV"].pop("drop_rules", None)
        else:
            cfgmod.CUSTOMER_OVERRIDES["MPAV"]["drop_rules"] = original


def test_summary_template(out_path):
    print("\n--- 场景二：【汇总】模板按标签写入 ---")
    ws = load_workbook(out_path)["汇总"]
    labels = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label and str(label) not in labels:      # 只取第一段，第二段标签会重名
            labels[str(label)] = r

    check("15. 账单初账户余额未被覆盖",
          ws.cell(labels["账单初账户余额"], 2).value == -1000.0,
          f"实际 {ws.cell(labels['账单初账户余额'], 2).value}")

    total_cell = ws.cell(labels["减 本期账单总支出"], 2).value
    check("15b. 本期总支出改为按费用行求和",
          isinstance(total_cell, str) and total_cell.startswith("=B"), f"实际 {total_cell}")

    out_cell = ws.cell(labels["出库订单费"], 2).value
    check("15c. 出库订单费链接到明细表",
          isinstance(out_cell, str) and "出库订单" in out_cell, f"实际 {out_cell}")

    rent_cell = ws.cell(labels["仓租费"], 2).value
    check("15d. 仓租费链接到仓租表",
          isinstance(rent_cell, str) and "仓租" in rent_cell, f"实际 {rent_cell}")

    closing = ws.cell(labels["账期末账号余额"], 2).value
    check("15e. 期末余额改为公式",
          isinstance(closing, str) and closing.startswith("="), f"实际 {closing}")

    # 第二段（物流账户）不应被破坏：模板里"客户编号"出现两次
    first_col = [str(ws.cell(r, 1).value) for r in range(1, ws.max_row + 1)]
    check("15f. 汇总第二段（物流账户）未被破坏", first_col.count("客户编号") == 2,
          f"实际出现 {first_col.count('客户编号')} 次")


def test_missing_orders(tmpdir):
    print("\n--- 场景三：账单原值 vs 明细合计（漏单检出）---")
    main_path = os.path.join(tmpdir, "gap.xlsx")
    out_path = os.path.join(tmpdir, "gap_已处理.xlsx")
    # 汇总里写一个比明细合计更大的金额，模拟 MPAV 的 613 元漏单
    build_main_file(main_path, outbound_total_in_summary=-99999.0)

    result = core.run_pipeline(
        main_path=main_path, year=2026, month=7, output_path=out_path,
        customer_code="MPAV", fuel_periods=PERIODS_JULY, log=lambda _t: None)

    flagged = [r for r in result["reconciliation"] if str(r["说明"]).startswith("★")]
    check("16. 检出账单金额与明细合计不符", any(r["项目"] == "出库订单费" for r in flagged),
          f"实际 {result['reconciliation']}")
    ws = load_workbook(out_path)["核对汇总"]
    texts = [str(ws.cell(r, 1).value) for r in range(1, ws.max_row + 1)]
    check("16b. 对账表已写入「核对汇总」", "【账单原值 vs 明细重算值 对账】" in texts)


def test_guxe(tmpdir):
    print("\n--- 场景四：GUXE（无燃油客户）---")
    main_path = os.path.join(tmpdir, "guxe.xlsx")
    out_path = os.path.join(tmpdir, "guxe_已处理.xlsx")
    build_guxe_file(main_path)

    result = core.run_pipeline(
        main_path=main_path, year=2026, month=7, output_path=out_path,
        customer_code="GUXE", log=lambda _t: None)

    check("8. 无燃油客户无需填写费率即可运行", os.path.exists(out_path))
    oc = read_sheet(out_path, "逐单核对")
    check("8b. 所有订单燃油费为 0", all(abs(float(v)) < 1e-9 for v in oc["实际燃油费"]))
    check("8c. 无燃油订单数量统计正确", result["stats"]["无燃油订单数量"] == len(oc))
    check("10c. GUXE 复合订单两行 SKU 均保留",
          list(oc["订单编号"].astype(str)).count("G002") == 2)
    check("8d. 重算结果与源账单完全一致",
          abs(result["stats"]["总计差额"]) < 1e-9, f"差额 {result['stats']['总计差额']}")


def test_invalid_date_block(tmpdir):
    print("\n--- 场景五：1970 年日期必须阻断 ---")
    main_path = os.path.join(tmpdir, "bad_date.xlsx")
    out_path = os.path.join(tmpdir, "bad_date_已处理.xlsx")

    rows = [order_row("C001", "SKU1", "卡派自提", datetime(2026, 7, 8), datetime(1970, 1, 1))]
    with pd.ExcelWriter(main_path, engine="openpyxl") as writer:
        summary_frame("TEST", 0).to_excel(writer, sheet_name="汇总", index=False)
        pd.DataFrame([{"时间": datetime(2026, 7, 1), "仓库": "W1", "实际费用": -1.0}]
                     ).to_excel(writer, sheet_name="仓租", index=False)
        pd.DataFrame(rows, columns=MAIN_COLUMNS).to_excel(writer, sheet_name="出库订单", index=False)

    try:
        core.run_pipeline(main_path=main_path, year=2026, month=7, output_path=out_path,
                          customer_code="MPAV", fuel_periods=PERIODS_JULY, log=lambda _t: None)
        check("13c. 卡派自提遇 1970 完成时间时阻断", False, "未抛出异常")
    except core.BillingProcessError as exc:
        check("13c. 卡派自提遇 1970 完成时间时阻断",
              "账期日期" in str(exc) or "1970" in str(exc), str(exc)[:120])


def test_structure_guard(tmpdir):
    print("\n--- 场景六：字段结构异常必须阻断 ---")
    main_path = os.path.join(tmpdir, "main2.xlsx")
    bad_path = os.path.join(tmpdir, "bad_supp.xlsx")
    out_path = os.path.join(tmpdir, "main2_已处理.xlsx")
    build_main_file(main_path)

    pd.DataFrame([{
        "订单编号": "B001", "SKU": "SKU1", "尾程物流": "平台快递", "订单类型": "单件代发",
        "创建时间": datetime(2026, 7, 8), "完成时间": datetime(2026, 7, 9),
        "订单处理费": -2.2, "基础运费": 10, "住宅地址附加费": 0,
        "新增未知字段": 1,
    }]).to_excel(bad_path, sheet_name="出库订单", index=False)

    try:
        core.run_pipeline(main_path=main_path, supplement_paths=[bad_path],
                          year=2026, month=7, output_path=out_path,
                          customer_code="MPAV", fuel_periods=PERIODS_JULY, log=lambda _t: None)
        check("1c. 补充文件出现多余字段时停止计算", False, "未抛出异常")
    except core.BillingProcessError as exc:
        check("1c. 补充文件出现多余字段时停止计算",
              "字段结构校验未通过" in str(exc), str(exc)[:100])


def test_rate_validation():
    print("\n--- 场景七：燃油费率区间校验 ---")
    overlap = [fr.FuelPeriod(date(2026, 7, 6), date(2026, 7, 12), Decimal("0.25")),
               fr.FuelPeriod(date(2026, 7, 10), date(2026, 7, 16), Decimal("0.26"))]
    check("17. 检测出费率区间重叠", any("重叠" in p for p in fr.validate_periods(overlap)))

    gap = [fr.FuelPeriod(date(2026, 7, 6), date(2026, 7, 12), Decimal("0.25")),
           fr.FuelPeriod(date(2026, 7, 20), date(2026, 7, 26), Decimal("0.26"))]
    check("17b. 检测出费率区间空档", any("空档" in p for p in fr.validate_periods(gap)))

    check("17c. 检测出未覆盖的订单日期",
          any("未被任何费率区间覆盖" in p
              for p in fr.validate_periods(PERIODS_JULY[:2], required_dates=[date(2026, 7, 25)])))

    check("9b. FedEx 官网周区间不做 ±1 天平移",
          fr.carrier_week_ranges(date(2026, 7, 1), date(2026, 7, 31))[0] ==
          (date(2026, 6, 29), date(2026, 7, 5)))


def test_rate_memory(tmpdir):
    print("\n--- 场景八：燃油费率记忆 ---")
    memory_file = os.path.join(tmpdir, "fuel_rate_memory.json")
    original = rate_store.store_path
    rate_store.store_path = lambda: memory_file
    try:
        rate_store.clear()
        check("18. 初始状态无记忆", rate_store.load_periods() == [])

        rate_store.remember(PERIODS_JULY[:2])
        loaded = rate_store.load_periods()
        check("18b. 费率写入后可读回", len(loaded) == 2, f"实际 {len(loaded)}")
        check("18c. 费率数值用 Decimal 精确保存",
              loaded[0].rate == Decimal("0.2525"), f"实际 {loaded[0].rate}")

        weeks = [(p.start, p.end) for p in PERIODS_JULY]
        missing = rate_store.missing_weeks(loaded, weeks, "FedEx")
        check("18d. 正确识别出还缺 3 个周", len(missing) == 3, f"实际 {len(missing)}")

        rate_store.remember(PERIODS_JULY[2:])
        check("18e. 补齐后不再缺周",
              rate_store.missing_weeks(rate_store.load_periods(), weeks, "FedEx") == [])

        # 同一区间重新记忆应覆盖而不是重复
        rate_store.remember([fr.FuelPeriod(date(2026, 7, 6), date(2026, 7, 12), Decimal("0.30"))])
        after = rate_store.load_periods()
        check("18f. 相同区间重复记忆会覆盖而非叠加", len(after) == 5, f"实际 {len(after)}")
        rate = next(p.rate for p in after if p.start == date(2026, 7, 6))
        check("18g. 覆盖后取到新费率", rate == Decimal("0.30"), f"实际 {rate}")

        rate_store.clear()
        check("18h. 清空记忆生效", rate_store.load_periods() == [])
    finally:
        rate_store.store_path = original


def test_rate_inference(tmpdir):
    print("\n--- 场景九：从源账单反推燃油费率 ---")
    main_path = os.path.join(tmpdir, "infer.xlsx")
    rows = []
    for day, rate in ((6, Decimal("0.25")), (8, Decimal("0.25")), (21, Decimal("0.2525"))):
        base = Decimal("12.5")
        rows.append(order_row(f"R{day}", "SKU1", "平台快递",
                              datetime(2026, 7, day), datetime(2026, 7, day + 1),
                              src_fuel=float(round_money(base * rate))))
    with pd.ExcelWriter(main_path, engine="openpyxl") as writer:
        pd.DataFrame(rows, columns=MAIN_COLUMNS).to_excel(writer, sheet_name="出库订单", index=False)

    inferred = {start: rate for start, _e, rate, n, _note
                in core.infer_rates_from_file(main_path, "MPAV", 2026, 7) if n}
    check("19. 反推出 7/6 那周约 25%",
          inferred.get(date(2026, 7, 6)) == Decimal("0.25"), f"实际 {inferred.get(date(2026, 7, 6))}")
    check("19b. 反推出 7/20 那周约 25.25%",
          inferred.get(date(2026, 7, 20)) == Decimal("0.2525"),
          f"实际 {inferred.get(date(2026, 7, 20))}")


def test_config_sanity():
    print("\n--- 场景十：客户配置自检 ---")
    for code in cfgmod.list_customer_codes():
        problems = cfgmod.validate_config(cfgmod.get_customer_config(code))
        check(f"配置校验通过：{code}", not problems, str(problems))

    cfg = cfgmod.get_customer_config("HBKE")
    check("账期字段：卡派自提 -> 完成时间",
          cfgmod.period_field_for(cfg, "卡派自提") == ("完成时间", False))
    check("账期字段：平台快递 -> 创建时间",
          cfgmod.period_field_for(cfg, "平台快递") == ("创建时间", False))
    check("账期字段：未知类型走兜底并告警",
          cfgmod.period_field_for(cfg, "海运整柜") == ("创建时间", True))
    check("GUXE 燃油模式为 none",
          cfgmod.get_customer_config("GUXE")["fuel"]["mode"] == "none")


def main():
    tmpdir = tempfile.mkdtemp(prefix="billing_test_")
    try:
        out_path = test_main(tmpdir)
        test_test_order_rule(tmpdir)
        test_summary_template(out_path)
        test_missing_orders(tmpdir)
        test_guxe(tmpdir)
        test_invalid_date_block(tmpdir)
        test_structure_guard(tmpdir)
        test_rate_validation()
        test_rate_memory(tmpdir)
        test_rate_inference(tmpdir)
        test_config_sanity()
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

    print("\n" + "=" * 62)
    print(f"通过 {len(PASSED)} 项，失败 {len(FAILED)} 项")
    for name, detail in FAILED:
        print(f"  [FAIL] {name} {detail}")
    print("=" * 62)
    return 1 if FAILED else 0


if __name__ == "__main__":
    sys.exit(main())
