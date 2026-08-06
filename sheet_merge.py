# -*- coding: utf-8 -*-
"""
sheet_merge.py —— 主文件 + 补充文件的读取、结构校验与合并

对应《代码修改建议》：
   1. 合并补充文件时**禁止按列位置拼接**，必须按表头名称映射
   2. 主文件和补充文件必须先合并再计算，并保留来源可追溯信息
  14. 复合订单不能按订单编号简单去重，业务主键 = 订单编号 + SKU

绝对禁止的写法（就是造成 MPAV 53 笔订单错位的那种）：
    target_sheet.append(source_row_values)

本模块的写法：
    target_row = {field: source_row.get(field) for field in standard_columns}
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field as dc_field
from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

# 追溯用的内部列，最终会写进输出文件，方便对账时逐笔回溯
COL_SOURCE_FILE = "__来源文件"
COL_SOURCE_SHEET = "__来源sheet"
COL_SOURCE_ROW = "__来源行号"
COL_SOURCE_ROLE = "__数据来源"          # 主文件 / 补充文件
COL_OVERRIDDEN = "__是否被补充覆盖"
COL_UPDATED_AT = "__更新时间"

PROVENANCE_COLUMNS = [
    COL_SOURCE_ROLE, COL_SOURCE_FILE, COL_SOURCE_SHEET,
    COL_SOURCE_ROW, COL_OVERRIDDEN, COL_UPDATED_AT,
]


class StructureError(Exception):
    """字段结构不一致，必须停止计算而不是继续运行（建议 1）。"""


@dataclass
class SourceSheet:
    path: str
    sheet_name: str
    role: str                      # "主文件" / "补充文件"
    headers: List[str]
    rows: List[Dict[str, Any]]

    @property
    def file_name(self) -> str:
        return os.path.basename(self.path)


@dataclass
class MergeResult:
    headers: List[str]
    rows: List[Dict[str, Any]]
    override_records: List[Dict[str, Any]] = dc_field(default_factory=list)
    duplicate_records: List[Dict[str, Any]] = dc_field(default_factory=list)
    added_from_supplement: int = 0
    structure_notes: List[str] = dc_field(default_factory=list)


# ---------------------------------------------------------------------------
# 表头规范化
# ---------------------------------------------------------------------------

def normalize_header(value: Any) -> str:
    """表头比对用的规范形式：去空白、去换行、全角转半角括号，忽略大小写差异。"""
    if value is None:
        return ""
    text = str(value)
    for ch in ("\n", "\r", "\t", " ", "\u3000"):
        text = text.replace(ch, "")
    text = text.replace("（", "(").replace("）", ")")
    return text.strip()


def build_header_index(headers: Sequence[str]) -> Dict[str, str]:
    """{规范化表头: 原始表头}"""
    index: Dict[str, str] = {}
    for header in headers:
        key = normalize_header(header)
        if key and key not in index:
            index[key] = header
    return index


def resolve_field(headers: Sequence[str], wanted: str) -> Optional[str]:
    """按表头文字找列（忽略空格/换行/全半角括号差异）。找不到返回 None。"""
    return build_header_index(headers).get(normalize_header(wanted))


def get_value(row: Dict[str, Any], headers: Sequence[str], wanted: str, default=None):
    actual = resolve_field(headers, wanted)
    if actual is None:
        return default
    return row.get(actual, default)


# ---------------------------------------------------------------------------
# 读取
# ---------------------------------------------------------------------------

def read_sheet(path: str, sheet_name: str, role: str = "主文件",
               header_row: int = 1) -> SourceSheet:
    """读取指定 sheet，返回表头 + 逐行字典（保留原始类型，日期仍是 datetime）。"""
    # 注意：这里刻意**不用** read_only 模式。
    # 本平台导出的账单里 dimension 记录是错的（写成 A1:A1），read_only 模式会据此
    # 把表头截断成"只有 1 列"，进而误报"找不到订单编号"。普通模式会重新扫描实际范围。
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as exc:
        raise StructureError(f"无法读取文件：{os.path.basename(path)}\n详细信息：{exc}") from exc

    try:
        target = _match_sheet_name(wb.sheetnames, sheet_name)
        if target is None:
            raise StructureError(
                f"文件【{os.path.basename(path)}】中找不到 sheet「{sheet_name}」。"
                f"现有 sheet：{'、'.join(wb.sheetnames)}"
            )

        ws = wb[target]
        rows_iter = ws.iter_rows(values_only=True)

        headers: List[str] = []
        for _ in range(header_row):
            try:
                raw_header = next(rows_iter)
            except StopIteration:
                raw_header = ()
        headers = [str(h).strip() if h is not None else "" for h in raw_header]

        # 处理同名字段重复（建议 1 的校验项之一）
        seen: Dict[str, int] = {}
        duplicated: List[str] = []
        final_headers: List[str] = []
        for h in headers:
            key = normalize_header(h)
            if key and key in seen:
                seen[key] += 1
                duplicated.append(h)
                final_headers.append(f"{h}#{seen[key]}")
            else:
                if key:
                    seen[key] = 1
                final_headers.append(h)

        if duplicated:
            raise StructureError(
                f"文件【{os.path.basename(path)}】的 sheet「{target}」存在同名字段重复："
                f"{'、'.join(sorted(set(duplicated)))}。请先在源文件中处理后再运行。"
            )

        rows: List[Dict[str, Any]] = []
        for row_no, values in enumerate(rows_iter, start=header_row + 1):
            if values is None or all(v is None or str(v).strip() == "" for v in values):
                continue
            row: Dict[str, Any] = {}
            for idx, header in enumerate(final_headers):
                if not header:
                    continue
                row[header] = values[idx] if idx < len(values) else None
            row[COL_SOURCE_FILE] = os.path.basename(path)
            row[COL_SOURCE_SHEET] = target
            row[COL_SOURCE_ROW] = row_no
            row[COL_SOURCE_ROLE] = role
            row[COL_OVERRIDDEN] = "否"
            row[COL_UPDATED_AT] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            rows.append(row)

        return SourceSheet(path=path, sheet_name=target, role=role,
                           headers=[h for h in final_headers if h], rows=rows)
    finally:
        wb.close()


def _match_sheet_name(sheet_names: Sequence[str], wanted: str) -> Optional[str]:
    target = normalize_header(wanted)
    for name in sheet_names:
        if normalize_header(name) == target:
            return name
    for name in sheet_names:
        if target and target in normalize_header(name):
            return name
    return None


# ---------------------------------------------------------------------------
# 结构校验（建议 1）
# ---------------------------------------------------------------------------

def compare_structure(main: SourceSheet, supplement: SourceSheet,
                      allow_extra_fields: bool = False) -> Tuple[List[str], List[str]]:
    """
    比对主文件与补充文件的字段结构。
    返回 (severe_problems, notes)。severe_problems 非空 = 必须停止计算。
    """
    severe: List[str] = []
    notes: List[str] = []

    main_index = build_header_index(main.headers)
    supp_index = build_header_index(supplement.headers)

    missing = [main_index[k] for k in main_index if k not in supp_index]
    extra = [supp_index[k] for k in supp_index if k not in main_index]

    if missing:
        notes.append(
            f"补充文件【{supplement.file_name}】缺少主文件的 {len(missing)} 个字段："
            f"{'、'.join(missing[:15])}{' 等' if len(missing) > 15 else ''}"
            f"（这些字段在合并后按空值处理，不会造成错位）。"
        )
    if extra:
        message = (
            f"补充文件【{supplement.file_name}】比主文件多出 {len(extra)} 个字段："
            f"{'、'.join(extra[:15])}{' 等' if len(extra) > 15 else ''}。"
        )
        if allow_extra_fields:
            notes.append(message + " 已按配置放行，这些字段会追加到输出表。")
        else:
            severe.append(
                message + " 请确认是否为表结构变更；确认无误后可在客户配置中打开 merge.allow_extra_fields。"
            )

    # 字段顺序变化：不影响本程序（按名映射），但要显式告知，便于发现源系统改版
    common = [k for k in main_index if k in supp_index]
    main_order = [k for k in (normalize_header(h) for h in main.headers) if k in supp_index]
    supp_order = [k for k in (normalize_header(h) for h in supplement.headers) if k in main_index]
    if main_order != supp_order and common:
        notes.append(
            f"补充文件【{supplement.file_name}】的字段顺序与主文件不同"
            f"（主文件 {len(main.headers)} 列 / 补充文件 {len(supplement.headers)} 列）。"
            f"本程序按表头名称映射，顺序不同不会错位。"
        )

    return severe, notes


# ---------------------------------------------------------------------------
# 合并（建议 2、14）
# ---------------------------------------------------------------------------

def _key_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def build_business_key(row: Dict[str, Any], headers: Sequence[str],
                       key_fields: Sequence[str]) -> Tuple[str, ...]:
    return tuple(_key_text(get_value(row, headers, f)) for f in key_fields)


def _business_payload(row: Dict[str, Any], headers: Sequence[str]) -> Tuple:
    """不含追溯列的业务内容，用于判断“完全一致的重复行”。"""
    return tuple(_key_text(row.get(h)) for h in headers)


def merge_sources(
    main: SourceSheet,
    supplements: Sequence[SourceSheet],
    key_fields: Sequence[str],
    supplement_priority: bool = True,
    allow_extra_fields: bool = False,
    strict_structure: bool = True,
    watch_fields: Optional[Sequence[str]] = None,
) -> MergeResult:
    """
    按表头名称把主文件与补充文件合并成一张表。

    * 主键相同 + 补充文件优先 → 覆盖，并记录到「补充文件覆盖记录」
    * 主键不存在 → 作为新增订单加入
    * 同一文件内主键重复：内容完全一致 → 记为重复订单并丢弃；内容不同 → 严重异常
    """
    severe_all: List[str] = []
    notes: List[str] = []

    # 输出表头 = 主文件表头 + （允许时）补充文件独有表头
    headers: List[str] = list(main.headers)
    header_keys = {normalize_header(h) for h in headers}

    for supp in supplements:
        severe, supp_notes = compare_structure(main, supp, allow_extra_fields=allow_extra_fields)
        severe_all.extend(severe)
        notes.extend(supp_notes)
        if allow_extra_fields:
            for h in supp.headers:
                if normalize_header(h) not in header_keys:
                    headers.append(h)
                    header_keys.add(normalize_header(h))

    if severe_all and strict_structure:
        raise StructureError(
            "字段结构校验未通过，已停止计算：\n  - " + "\n  - ".join(severe_all)
        )
    notes.extend(severe_all)

    # 主键字段必须存在
    usable_keys = [f for f in key_fields if resolve_field(headers, f)]
    if not usable_keys:
        raise StructureError(
            f"业务主键字段 {'、'.join(key_fields)} 在表头中都找不到，无法安全合并。"
            f"请核对客户配置 merge.key 与实际表头。"
        )
    if len(usable_keys) < len(key_fields):
        missing_keys = [f for f in key_fields if f not in usable_keys]
        notes.append(
            f"业务主键字段 {'、'.join(missing_keys)} 在表中不存在，本次按 "
            f"{'+'.join(usable_keys)} 作为主键。"
        )

    watch = list(watch_fields or [])

    merged: Dict[Tuple[str, ...], Dict[str, Any]] = {}
    order: List[Tuple[str, ...]] = []
    duplicate_records: List[Dict[str, Any]] = []
    override_records: List[Dict[str, Any]] = []
    added_from_supplement = 0

    def standardize(row: Dict[str, Any]) -> Dict[str, Any]:
        """★ 按字段名映射，绝不按列位置拼接。"""
        target = {field: row.get(field) for field in headers}
        for col in PROVENANCE_COLUMNS:
            target[col] = row.get(col)
        return target

    def register(source: SourceSheet, row: Dict[str, Any]):
        nonlocal added_from_supplement
        key = build_business_key(row, source.headers, usable_keys)
        std = standardize(row)

        if key not in merged:
            merged[key] = std
            order.append(key)
            if source.role == "补充文件":
                added_from_supplement += 1
            return

        existing = merged[key]
        same_file = existing.get(COL_SOURCE_FILE) == std.get(COL_SOURCE_FILE)
        identical = _business_payload(existing, headers) == _business_payload(std, headers)

        if identical:
            duplicate_records.append({
                "业务主键": " | ".join(key),
                "来源文件": std.get(COL_SOURCE_FILE),
                "来源行号": std.get(COL_SOURCE_ROW),
                "保留行": f"{existing.get(COL_SOURCE_FILE)} 第 {existing.get(COL_SOURCE_ROW)} 行",
                "判定": "整行内容完全一致的重复记录，已丢弃",
            })
            return

        if same_file:
            duplicate_records.append({
                "业务主键": " | ".join(key),
                "来源文件": std.get(COL_SOURCE_FILE),
                "来源行号": std.get(COL_SOURCE_ROW),
                "保留行": f"{existing.get(COL_SOURCE_FILE)} 第 {existing.get(COL_SOURCE_ROW)} 行",
                "判定": "★同一文件内业务主键重复但内容不同，请人工确认（已保留先出现的一行）",
            })
            return

        if source.role == "补充文件" and supplement_priority:
            changes = []
            for f in watch:
                old = get_value(existing, headers, f)
                new = get_value(std, headers, f)
                if _key_text(old) != _key_text(new):
                    changes.append(f"{f}: {_key_text(old)} -> {_key_text(new)}")
            std[COL_OVERRIDDEN] = "是"
            override_records.append({
                "业务主键": " | ".join(key),
                "被覆盖来源": f"{existing.get(COL_SOURCE_FILE)} 第 {existing.get(COL_SOURCE_ROW)} 行",
                "覆盖来源": f"{std.get(COL_SOURCE_FILE)} 第 {std.get(COL_SOURCE_ROW)} 行",
                "变化字段": "；".join(changes) if changes else "（关注字段无变化，整行以补充文件为准）",
                "更新时间": std.get(COL_UPDATED_AT),
            })
            merged[key] = std
        else:
            override_records.append({
                "业务主键": " | ".join(key),
                "被覆盖来源": "（未覆盖）",
                "覆盖来源": f"{std.get(COL_SOURCE_FILE)} 第 {std.get(COL_SOURCE_ROW)} 行",
                "变化字段": "按配置以主文件为准，补充文件该行被忽略",
                "更新时间": std.get(COL_UPDATED_AT),
            })

    for row in main.rows:
        register(main, row)
    for supp in supplements:
        for row in supp.rows:
            register(supp, row)

    return MergeResult(
        headers=headers + PROVENANCE_COLUMNS,
        rows=[merged[k] for k in order],
        override_records=override_records,
        duplicate_records=duplicate_records,
        added_from_supplement=added_from_supplement,
        structure_notes=notes,
    )
