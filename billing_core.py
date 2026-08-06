# -*- coding: utf-8 -*-
"""
billing_core.py —— 账单处理核心逻辑（v4，按《代码修改建议》重构）

与旧版（legacy_billing_core_v3.py）的关键差异
--------------------------------------------
1. 多文件输入：主文件 + N 个补充文件，按**表头名称**映射合并，绝不按列位置拼接。
2. 先合并再计算，保留来源文件 / 行号 / 是否被覆盖，可逐笔追溯。
3. 不再因“订单处理费为 0 或为空”删除订单，只记异常。
4. 账期归属字段由**物流类型**决定（卡派自提 -> 完成时间，其余 -> 创建时间）。
5. 费用总计按客户配置的字段清单汇总，不再依赖 AI/AL/AY 等列字母。
6. 燃油、物流费全部按表头名定位；PlatformFee 纳入物流费用。
7. 燃油支持 none / schedule / source 三种模式，无燃油客户可直接运行。
8. 金额全程 Decimal，逐单四舍五入（ROUND_HALF_UP）后再汇总。
9. 输出固定包含核账与异常明细工作表。
"""

from __future__ import annotations

import calendar
import os
import re
import tempfile
from dataclasses import dataclass, field as dc_field
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import customer_config as cfgmod
import fuel_rates as fr
import sheet_merge as sm
from money import (
    ZERO, MoneyParseError, decimal_to_float, format_rate, is_blank_value,
    looks_like_money_text, parse_rate, round_money, to_decimal,
)

# ------------------------- 基础配置 -------------------------

SHEET_WAREHOUSE_RENT = "仓租"
SHEET_INBOUND_DETAIL = "入库详情"
SHEET_INBOUND_ORDER = "入库订单"
SHEET_OUTBOUND_ORDER = "出库订单"
SHEET_FUEL_RATE = "燃油费率表"
SHEET_OTHER_SUMMARY = "其他费用汇总"
DEFAULT_MAIN_SUMMARY_SHEET = "汇总"

# 固定输出的核账 / 异常工作表（建议 22）
SHEET_CHECK_SUMMARY = "核对汇总"
SHEET_ORDER_CHECK = "逐单核对"
SHEET_MISSING_ORDERS = "漏单明细"
SHEET_DUPLICATE_ORDERS = "重复订单"
SHEET_CROSS_MONTH = "跨月订单"
SHEET_OVERRIDE_LOG = "补充文件覆盖记录"
SHEET_RATE_ISSUES = "费率异常"
SHEET_FIELD_ISSUES = "费用字段异常"
SHEET_UNEXPLAINED = "无法解释费用"
SHEET_MANUAL_ADJUST = "手工调整"
SHEET_DATA_QUALITY = "数据质量报告"

REPORT_SHEETS = [
    SHEET_CHECK_SUMMARY, SHEET_ORDER_CHECK, SHEET_MISSING_ORDERS, SHEET_DUPLICATE_ORDERS,
    SHEET_CROSS_MONTH, SHEET_OVERRIDE_LOG, SHEET_RATE_ISSUES, SHEET_FIELD_ISSUES,
    SHEET_UNEXPLAINED, SHEET_MANUAL_ADJUST, SHEET_DATA_QUALITY,
]

# 核算过程列，写进【出库订单】便于人工复核
COL_PERIOD_FIELD = "核算_账期字段"
COL_PERIOD_DATE = "核算_账期日期"
COL_FUEL_BASE = "核算_燃油基数"
COL_FUEL_RATE = "核算_燃油费率"
AUDIT_COLUMNS = [COL_PERIOD_FIELD, COL_PERIOD_DATE, COL_FUEL_BASE, COL_FUEL_RATE]

RENT_DATE_COL = "时间"

RENT_AMOUNT_HEADER = "实际费用"
RENT_AMOUNT_FALLBACK_LETTER = "R"
INBOUND_AMOUNT_HEADER = "费用"
INBOUND_AMOUNT_FALLBACK_LETTER = "K"

SHEETS_ALREADY_COUNTED = {
    SHEET_INBOUND_ORDER: (
        SHEET_INBOUND_DETAIL,
        "与【入库详情】为同一笔钱（汇总表 vs 明细表），已由【入库详情】计入，跳过以避免重复计算。",
    ),
}

# 固定 sheet 的汇总规则：只允许汇总指定表头，避免误算其他金额列。
# display_name 必须与【汇总】sheet 模板里 A 列的标签文字一致，程序按标签定位写入行。
FIXED_OTHER_SHEET_RULES = {
    "工单": {"target_header": "费用总计", "display_name": "工单处理费"},
    "退件订单": {"target_header": "费用总计", "display_name": "退件处理费"},
    "核账补收": {"target_header": "补收费用", "display_name": "核账补收费"},
    "询价打单": {"target_header": "总费用", "display_name": "询价打单费"},
    "其他": {"target_header": "实际到账", "display_name": "其他费"},
    "充值入账": {"target_header": "实际到账", "display_name": "加 账单充值总额"},
    "退款详情": {"target_header": "费用", "display_name": "退款/返款"},
    # 兼容旧版账单里可能出现的表
    "索赔抵扣": {"target_header": "索赔结果", "display_name": "索赔抵扣"},
}

UNKNOWN_AMOUNT_KEYWORDS = [
    "费用总计", "总费用", "实际费用", "费用", "金额", "收费金额", "账单金额",
    "应收金额", "应付金额", "服务费", "操作费", "运费", "补收", "退款", "赔偿", "索赔",
]

# ---- 【汇总】sheet 是平台给的固定模板，A 列是标签、B 列是金额 ----
# 真实结构（以 MPAV 为例）：
#   1  类别 / 金额
#   2-6  客户编号、客户昵称、公司名称、账单币种、账单起止日期
#   7  账单初账户余额
#   8  减 本期账单总支出      = SUM(第9~16行)
#   9-16 出库订单费、入库订单费、工单处理费、退件处理费、仓租费、询价打单费、核账补收费、其他费
#   17 加 账单充值总额
#   18 退款/返款
#   19 账期末账号余额          = B7 + B8 + B17 + B18
#   20 空行
#   21-34 物流账户部分（本程序不改动）
# ★ 因此绝不能像旧版那样"从第7行起覆盖写"，必须**按标签文字定位行**。
SUMMARY_LABEL_OPENING_BALANCE = "账单初账户余额"
SUMMARY_LABEL_TOTAL_EXPENSE = "减 本期账单总支出"
SUMMARY_LABEL_RECHARGE = "加 账单充值总额"
SUMMARY_LABEL_REFUND = "退款/返款"
SUMMARY_LABEL_CLOSING_BALANCE = "账期末账号余额"

# 计入「减 本期账单总支出」的费用标签，顺序与模板一致
SUMMARY_EXPENSE_LABELS = [
    "出库订单费", "入库订单费", "工单处理费", "退件处理费",
    "仓租费", "询价打单费", "核账补收费", "其他费",
]

MAIN_SUMMARY_DISPLAY_ORDER = SUMMARY_EXPENSE_LABELS + [
    SUMMARY_LABEL_RECHARGE, SUMMARY_LABEL_REFUND,
]

PROTECTED_SHEETS_FROM_REMOVAL = {
    DEFAULT_MAIN_SUMMARY_SHEET, SHEET_WAREHOUSE_RENT, SHEET_OUTBOUND_ORDER,
    SHEET_FUEL_RATE, SHEET_OTHER_SUMMARY, *REPORT_SHEETS,
}

LogFunc = Callable[[str], None]


def _noop_log(_text: str) -> None:
    return None


# ------------------------- 异常类型 -------------------------

class BillingProcessError(Exception):
    """可预期的处理错误，GUI 以友好文案展示。"""


class PipelineCancelled(Exception):
    """用户在交互环节主动取消。"""


class OrderCountMismatchError(Exception):
    """订单件数核对不一致且用户选择不继续。"""


class PrecheckBlocked(BillingProcessError):
    """运行前检查发现严重异常，禁止继续计算（建议 十一）。"""


# ------------------------- 异常收集 -------------------------

SEVERITY_BLOCK = "严重"
SEVERITY_WARN = "提示"


@dataclass
class Issue:
    category: str
    severity: str
    order_id: str = ""
    sku: str = ""
    source_file: str = ""
    source_row: Any = ""
    issue_type: str = ""
    source_amount: Any = ""
    recalc_amount: Any = ""
    diff: Any = ""
    reason: str = ""
    suggestion: str = ""

    def to_row(self, customer_code: str) -> Dict[str, Any]:
        return {
            "客户编码": customer_code,
            "订单编号": self.order_id,
            "SKU": self.sku,
            "源文件": self.source_file,
            "源文件行号": self.source_row,
            "异常类型": self.issue_type or self.category,
            "严重程度": self.severity,
            "源金额": self.source_amount,
            "重算金额": self.recalc_amount,
            "差额": self.diff,
            "原因": self.reason,
            "建议处理": self.suggestion,
        }


class IssueCollector:
    def __init__(self, customer_code: str):
        self.customer_code = customer_code
        self.issues: List[Issue] = []

    def add(self, **kwargs) -> None:
        self.issues.append(Issue(**kwargs))

    def by_category(self, category: str) -> List[Dict[str, Any]]:
        return [i.to_row(self.customer_code) for i in self.issues if i.category == category]

    def blocking(self) -> List[Issue]:
        return [i for i in self.issues if i.severity == SEVERITY_BLOCK]

    def count(self, category: Optional[str] = None) -> int:
        if category is None:
            return len(self.issues)
        return sum(1 for i in self.issues if i.category == category)


# ------------------------- 通用工具 -------------------------

def norm_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("\n", "").replace("\r", "").replace(" ", "")


def quote_sheet_name(sheet_name: str) -> str:
    return "'" + sheet_name.replace("'", "''") + "'"


def is_blank(value) -> bool:
    return value is None or str(value).strip() == ""


def is_summary_or_archive_sheet(sheet_name: str, main_summary_sheet: str) -> bool:
    if sheet_name in {main_summary_sheet, SHEET_FUEL_RATE, SHEET_OTHER_SUMMARY}:
        return True
    if sheet_name in REPORT_SHEETS:
        return True
    archive_suffixes = [
        "_重复数据", "_月份外已删除", "_AI列为空已删除", "_已删除", "重复数据", "月份外已删除",
    ]
    return any(sheet_name.endswith(suffix) for suffix in archive_suffixes)


def find_header_col(ws: Worksheet, header_name: str, header_row: int = 1) -> Optional[int]:
    target = sm.normalize_header(header_name)
    for cell in ws[header_row]:
        if sm.normalize_header(cell.value) == target:
            return cell.column
    return None


def find_header_col_by_keywords(ws: Worksheet, keywords: Sequence[str], header_row: int = 1) -> List[Tuple[int, str]]:
    candidates = []
    for cell in ws[header_row]:
        header = norm_text(cell.value)
        if not header:
            continue
        if any(keyword in header for keyword in keywords):
            candidates.append((cell.column, header))
    return candidates


def find_last_data_row(ws: Worksheet, key_cols: Optional[Sequence[int]] = None, start_row: int = 2) -> int:
    total_labels = {"合计", "总计", "费用合计", "费用总计", "实际费用合计", "索赔结果合计", "补收金额合计"}
    max_row = ws.max_row

    for row in range(max_row, start_row - 1, -1):
        row_values = [norm_text(ws.cell(row, col).value) for col in range(1, min(ws.max_column, 8) + 1)]
        if any(v in total_labels or v.endswith("合计") for v in row_values):
            continue

        if key_cols:
            for col in key_cols:
                if col <= ws.max_column and not is_blank(ws.cell(row, col).value):
                    return row
        else:
            for col in range(1, ws.max_column + 1):
                if not is_blank(ws.cell(row, col).value):
                    return row

    return start_row - 1


def remove_existing_total_rows(ws: Worksheet, amount_col: int, label_texts: Sequence[str], header_row: int = 1) -> None:
    label_col = max(1, amount_col - 1)
    labels = {norm_text(x) for x in label_texts}
    initial_max_row = ws.max_row
    tail_start_row = max(header_row + 1, initial_max_row - 5)

    for row in range(initial_max_row, header_row, -1):
        label_value = norm_text(ws.cell(row, label_col).value)
        amount_formula = norm_text(ws.cell(row, amount_col).value)

        if label_value in labels:
            ws.delete_rows(row, 1)
            continue

        a_value = norm_text(ws.cell(row, 1).value)
        if a_value in labels:
            ws.delete_rows(row, 1)
            continue

        if (
            row >= tail_start_row
            and (amount_formula.startswith("=SUM(") or amount_formula.startswith("=SUBTOTAL("))
            and ("合计" in label_value or "总计" in label_value or "合计" in a_value or "总计" in a_value)
        ):
            ws.delete_rows(row, 1)


def make_sum_formula(col_letter: str, start_row: int, end_row: int, formula_function: str = "SUM") -> str:
    if end_row < start_row:
        return "=0"
    if formula_function.upper() == "SUBTOTAL_109":
        return f"=SUBTOTAL(109,{col_letter}{start_row}:{col_letter}{end_row})"
    return f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"


def write_total_formula(
    ws: Worksheet,
    amount_col: int,
    total_label: str,
    formula_function: str = "SUM",
    start_row: int = 2,
    key_cols: Optional[Sequence[int]] = None,
) -> str:
    remove_existing_total_rows(ws, amount_col, [total_label])
    amount_letter = get_column_letter(amount_col)
    last_data_row = find_last_data_row(ws, key_cols=key_cols or [amount_col], start_row=start_row)
    total_row = max(start_row, last_data_row + 1)

    label_col = max(1, amount_col - 1)
    ws.cell(total_row, label_col).value = total_label
    ws.cell(total_row, amount_col).value = make_sum_formula(amount_letter, start_row, last_data_row, formula_function)
    ws.cell(total_row, label_col).font = Font(bold=True)
    ws.cell(total_row, amount_col).font = Font(bold=True)
    ws.cell(total_row, amount_col).number_format = '#,##0.00;[Red]-#,##0.00'
    return f"{amount_letter}{total_row}"


def coerce_excel_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        try:
            number = float(value)
        except (ValueError, OverflowError):
            return None
        if number != number:          # NaN
            return None
        if 1 <= number <= 60000:      # Excel 日期序列号
            return (datetime(1899, 12, 30) + timedelta(days=number)).date()
        return None

    text = str(value).strip()
    if not text:
        return None

    known_formats = (
        "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d",
        "%m/%d/%Y %H:%M:%S", "%m/%d/%Y", "%Y年%m月%d日",
    )
    for fmt in known_formats:
        try:
            return datetime.strptime(text[:19], fmt).date()
        except ValueError:
            pass

    try:
        from dateutil.parser import parse
        return parse(text).date()
    except Exception:
        return None


@dataclass
class FeeSummaryItem:
    display_name: str
    sheet_name: str
    total_cell: str
    amount_header: str = ""
    note: str = ""


# ------------------------- 出库订单核算引擎 -------------------------

@dataclass
class OutboundResult:
    headers: List[str]
    rows: List[Dict[str, Any]]
    order_check_rows: List[Dict[str, Any]] = dc_field(default_factory=list)
    cross_month_rows: List[Dict[str, Any]] = dc_field(default_factory=list)
    dropped_rows: List[Dict[str, Any]] = dc_field(default_factory=list)
    stats: Dict[str, Any] = dc_field(default_factory=dict)


def _sum_fields(row: Dict[str, Any], headers: Sequence[str], fields: Sequence[str]) -> Decimal:
    """按表头名求和；配置里有、表里没有的字段按 0 处理（调用前已单独告警）。"""
    total = ZERO
    for name in fields:
        actual = sm.resolve_field(headers, name)
        if actual is None:
            continue
        total += to_decimal(row.get(actual), field=name)
    return total


def _is_test_order(row: Dict[str, Any], headers: Sequence[str], keywords: Sequence[str],
                   order_field: str) -> bool:
    if not keywords:
        return False
    watch_headers = {sm.normalize_header(order_field), "备注", "订单备注", "标签"}
    haystack = " ".join(
        str(row.get(h) or "") for h in headers if sm.normalize_header(h) in watch_headers
    )
    return any(kw and kw in haystack for kw in keywords)


def _audit_record(row, order_id, sku, src_file, src_row,
                  logistics_type, period_field, period_date) -> Dict[str, Any]:
    return {
        "订单编号": order_id,
        "SKU": sku,
        "物流类型": logistics_type,
        "账期字段": period_field,
        "账期日期": period_date.strftime("%Y-%m-%d") if period_date else "",
        "源文件": src_file,
        "源文件行号": src_row,
        "数据来源": row.get(sm.COL_SOURCE_ROLE, ""),
    }


def calculate_outbound(
    merge_result: sm.MergeResult,
    config: Dict[str, Any],
    year: int,
    month: int,
    periods: Sequence[fr.FuelPeriod],
    collector: IssueCollector,
    log: LogFunc = _noop_log,
) -> OutboundResult:
    """对合并后的出库订单逐单核算。金额全程 Decimal，逐单取整后再汇总。"""
    original_headers = list(merge_result.headers)
    headers = list(merge_result.headers)
    fields = config["fields"]
    rounding = config["rounding"]
    digits, rmode = rounding["digits"], rounding["mode"]
    fuel_cfg = config["fuel"]
    fuel_mode = fuel_cfg["mode"]
    no_fuel_logistics = {sm.normalize_header(x) for x in fuel_cfg.get("no_fuel_logistics") or []}
    drop_rules = config["drop_rules"]

    order_field = fields["order_id"]
    sku_field = fields["sku"]
    logistics_type_field = fields["logistics_type"]
    fuel_fee_field = fields["fuel_fee"]
    logistics_fee_field = fields["logistics_fee"]
    total_field = fields["total"]
    source_fuel_field = fields["source_fuel_fee"]

    # 结果列按“表头名”确保存在；不存在就新增一列，不会占用别的列
    for name in (fuel_fee_field, logistics_fee_field, total_field, *AUDIT_COLUMNS):
        if sm.resolve_field(headers, name) is None:
            headers.append(name)

    fuel_col = sm.resolve_field(headers, fuel_fee_field)
    logi_col = sm.resolve_field(headers, logistics_fee_field)
    total_col = sm.resolve_field(headers, total_field)
    source_total_col = sm.resolve_field(original_headers, total_field)

    # 费用总计构成里除物流费用之外的字段（物流费用用重算值）
    total_component_fields = [
        f for f in config["total_fields"]
        if sm.normalize_header(f) != sm.normalize_header(logistics_fee_field)
    ]
    present_total_fields = [f for f in total_component_fields if sm.resolve_field(original_headers, f)]
    absent_total_fields = [f for f in total_component_fields if not sm.resolve_field(original_headers, f)]
    if absent_total_fields:
        log(f"  [提示] 费用总计配置中以下字段本表不存在，已按 0 处理：{'、'.join(absent_total_fields)}")

    present_base_fields = [f for f in config["fuel_base_fields"] if sm.resolve_field(original_headers, f)]
    absent_base_fields = [f for f in config["fuel_base_fields"] if not sm.resolve_field(original_headers, f)]
    present_platform_fields = [f for f in config["platform_fee_fields"] if sm.resolve_field(original_headers, f)]

    log(f"  燃油基数字段：{'、'.join(present_base_fields) or '（无）'}"
        + (f"｜配置中不存在的：{'、'.join(absent_base_fields)}" if absent_base_fields else ""))
    log(f"  PlatformFee 字段：{'、'.join(present_platform_fields) or '（本表无，按 0 计，公式项仍保留）'}")
    log(f"  费用总计构成：{'、'.join(present_total_fields) or '（无）'} + {logistics_fee_field}")

    kept_rows: List[Dict[str, Any]] = []
    order_check_rows: List[Dict[str, Any]] = []
    cross_month_rows: List[Dict[str, Any]] = []
    dropped_rows: List[Dict[str, Any]] = []

    logistics_type_counter: Dict[str, int] = {}
    fuel_order_count = 0
    no_fuel_order_count = 0
    period_fallback_count = 0
    zero_fee_count = 0
    invalid_date_count = 0
    seen_order_ids: Dict[str, int] = {}

    total_recalc = ZERO
    total_source = ZERO

    proc_field_actual = sm.resolve_field(original_headers, "订单处理费")
    status_field_actual = sm.resolve_field(original_headers, fields.get("status", "订单状态"))

    for row in merge_result.rows:
        order_id = str(sm.get_value(row, original_headers, order_field) or "").strip()
        sku = str(sm.get_value(row, original_headers, sku_field) or "").strip()
        src_file = row.get(sm.COL_SOURCE_FILE, "")
        src_row = row.get(sm.COL_SOURCE_ROW, "")
        logistics_type = str(sm.get_value(row, original_headers, logistics_type_field) or "").strip()
        key = logistics_type or "（空）"
        logistics_type_counter[key] = logistics_type_counter.get(key, 0) + 1

        # ---- 4. 账期字段由物流类型决定 ----
        period_field, fallback = cfgmod.period_field_for(config, logistics_type)
        if fallback:
            period_fallback_count += 1
            if config.get("warn_on_period_fallback", True) and logistics_type:
                collector.add(
                    category=SHEET_FIELD_ISSUES, severity=SEVERITY_WARN,
                    order_id=order_id, sku=sku, source_file=src_file, source_row=src_row,
                    issue_type="物流类型未配置账期规则",
                    reason=f"物流类型「{logistics_type}」不在 period_rules 中，已按默认字段【{period_field}】判断账期。",
                    suggestion="在 customer_config.py 的 period_rules 里补充该物流类型。",
                )

        raw_period_value = sm.get_value(row, original_headers, period_field)
        period_date = coerce_excel_date(raw_period_value)

        # ---- 3. 合并完成后仍出现 1970 年 / 无法解析的日期 -> 报错 ----
        if period_date is None or period_date.year < 2000:
            invalid_date_count += 1
            collector.add(
                category=SHEET_FIELD_ISSUES, severity=SEVERITY_BLOCK,
                order_id=order_id, sku=sku, source_file=src_file, source_row=src_row,
                issue_type="账期日期异常",
                source_amount=str(raw_period_value),
                reason=(f"字段【{period_field}】的值「{raw_period_value}」无法识别为有效日期"
                        + ("（1970 年通常是系统默认空值）" if period_date and period_date.year < 2000 else "")),
                suggestion="在源文件中修正该订单的时间字段，或确认该物流类型是否应改用其他账期字段。",
            )
            continue

        in_period = (period_date.year == year and period_date.month == month)

        status_text = str(row.get(status_field_actual) or "").strip() if status_field_actual else ""

        # ---- 3. 只有明确规则才允许剔除 ----
        if not in_period and drop_rules.get("drop_out_of_period", True):
            record = _audit_record(row, order_id, sku, src_file, src_row,
                                   logistics_type, period_field, period_date)
            record["剔除原因"] = f"账期外订单（{period_field}={period_date:%Y-%m-%d}，不属于 {year}-{month:02d}）"
            cross_month_rows.append(record)
            continue

        if status_text and any(s and s in status_text for s in drop_rules.get("cancelled_status") or []):
            record = _audit_record(row, order_id, sku, src_file, src_row,
                                   logistics_type, period_field, period_date)
            record["剔除原因"] = f"订单状态为「{status_text}」，属于已取消/作废订单"
            dropped_rows.append(record)
            continue

        if _is_test_order(row, original_headers, drop_rules.get("test_order_keywords") or [], order_field):
            record = _audit_record(row, order_id, sku, src_file, src_row,
                                   logistics_type, period_field, period_date)
            record["剔除原因"] = "已确认的测试订单"
            dropped_rows.append(record)
            continue

        # ---- 14. 复合订单：SKU 明细行全部保留，只标记是否为订单首行 ----
        first_row_of_order = order_id not in seen_order_ids
        seen_order_ids[order_id] = seen_order_ids.get(order_id, 0) + 1

        # ---- 金额（全程 Decimal） ----
        try:
            base_fee = _sum_fields(row, original_headers, present_base_fields)
            platform_fee = _sum_fields(row, original_headers, present_platform_fields)
            processing_total = _sum_fields(row, original_headers, present_total_fields)
        except MoneyParseError as exc:
            collector.add(
                category=SHEET_FIELD_ISSUES, severity=SEVERITY_BLOCK,
                order_id=order_id, sku=sku, source_file=src_file, source_row=src_row,
                issue_type="金额列出现文本", reason=str(exc),
                suggestion="在源文件中把该单元格改为纯数字或留空。",
            )
            continue

        # ---- 8. 燃油 ----
        applied_rate: Optional[Decimal] = None
        if fuel_mode == "none" or sm.normalize_header(logistics_type) in no_fuel_logistics:
            fuel_fee = ZERO
            applied_rate = ZERO
            no_fuel_order_count += 1
        elif fuel_mode == "source":
            fuel_fee = round_money(
                to_decimal(sm.get_value(row, original_headers, source_fuel_field), field=source_fuel_field),
                digits, rmode)
            fuel_order_count += 1
        else:  # schedule
            rate_field = fuel_cfg.get("rate_date_field", "period")
            if rate_field in ("period", "", None):
                fuel_date = period_date
            else:
                fuel_date = coerce_excel_date(sm.get_value(row, original_headers, rate_field)) or period_date
            applied_rate = fr.find_rate(periods, fuel_date, carrier=fuel_cfg.get("carrier"))
            if applied_rate is None:
                blocking = fuel_cfg.get("missing_rate_action", "error") == "error"
                collector.add(
                    category=SHEET_RATE_ISSUES,
                    severity=SEVERITY_BLOCK if blocking else SEVERITY_WARN,
                    order_id=order_id, sku=sku, source_file=src_file, source_row=src_row,
                    issue_type="燃油费率未覆盖该日期",
                    reason=f"日期 {fuel_date:%Y-%m-%d} 不在任何燃油费率区间内。",
                    suggestion="补齐燃油费率表对应区间后重新运行。",
                )
                if blocking:
                    continue
                applied_rate = ZERO
            fuel_fee = round_money(base_fee * applied_rate, digits, rmode)
            fuel_order_count += 1

        # ---- 7. 物流费用必须含 PlatformFee ----
        logistics_fee = -round_money(base_fee + fuel_fee + platform_fee, digits, rmode)

        # ---- 5/13. 费用总计 ----
        order_total = round_money(processing_total + logistics_fee, digits, rmode)

        # ---- 3. 订单处理费为 0/空：只提示，不删除 ----
        if proc_field_actual is not None:
            proc_value = row.get(proc_field_actual)
            if is_blank_value(proc_value) or to_decimal(proc_value, strict=False) == ZERO:
                zero_fee_count += 1
                collector.add(
                    category=SHEET_FIELD_ISSUES, severity=SEVERITY_WARN,
                    order_id=order_id, sku=sku, source_file=src_file, source_row=src_row,
                    issue_type="订单处理费为空或为0",
                    source_amount=str(proc_value),
                    recalc_amount=decimal_to_float(order_total),
                    reason="订单处理费为空或为 0，但订单属于本账期、物流费仍然有效，已保留计费。",
                    suggestion="确认是否应免收处理费；若为复合订单的 SKU 明细行属正常现象。",
                )

        # ---- 与源文件费用总计对比 ----
        source_total_raw = row.get(source_total_col) if source_total_col else None
        source_total = None
        if not is_blank_value(source_total_raw):
            source_total = round_money(to_decimal(source_total_raw, strict=False), digits, rmode)
        diff = (order_total - source_total) if source_total is not None else None

        if diff is not None and diff != ZERO:
            collector.add(
                category=SHEET_UNEXPLAINED, severity=SEVERITY_WARN,
                order_id=order_id, sku=sku, source_file=src_file, source_row=src_row,
                issue_type="重算总计与源文件不一致",
                source_amount=decimal_to_float(source_total),
                recalc_amount=decimal_to_float(order_total),
                diff=decimal_to_float(diff),
                reason="源文件费用总计与按配置重算的结果存在差额。",
                suggestion="核对 total_fields 字段清单是否完整、燃油费率是否正确。",
            )

        # ---- 写回结果 ----
        row[fuel_col] = decimal_to_float(fuel_fee)
        row[logi_col] = decimal_to_float(logistics_fee)
        row[total_col] = decimal_to_float(order_total)
        row[COL_PERIOD_FIELD] = period_field
        row[COL_PERIOD_DATE] = period_date.strftime("%Y-%m-%d")
        row[COL_FUEL_BASE] = decimal_to_float(round_money(base_fee, digits, rmode))
        row[COL_FUEL_RATE] = decimal_to_float(applied_rate) if applied_rate is not None else ""

        kept_rows.append(row)
        total_recalc += order_total
        if source_total is not None:
            total_source += source_total

        order_check_rows.append({
            "客户编码": config["customer_code"],
            "订单编号": order_id,
            "SKU": sku,
            "物流类型": logistics_type,
            "账期字段": period_field,
            "账期日期": period_date.strftime("%Y-%m-%d"),
            "是否订单首行": "是" if first_row_of_order else "否（SKU明细）",
            "处理费类合计": decimal_to_float(round_money(processing_total, digits, rmode)),
            "燃油基数": decimal_to_float(round_money(base_fee, digits, rmode)),
            "燃油费率": format_rate(applied_rate) if applied_rate is not None else "（源文件金额）",
            "实际燃油费": decimal_to_float(fuel_fee),
            "PlatformFee": decimal_to_float(round_money(platform_fee, digits, rmode)),
            "物流费用": decimal_to_float(logistics_fee),
            "重算费用总计": decimal_to_float(order_total),
            "源文件费用总计": decimal_to_float(source_total) if source_total is not None else "",
            "差额": decimal_to_float(diff) if diff is not None else "",
            "数据来源": row.get(sm.COL_SOURCE_ROLE, ""),
            "源文件": src_file,
            "源文件行号": src_row,
            "是否被补充覆盖": row.get(sm.COL_OVERRIDDEN, ""),
        })

    if invalid_date_count:
        log(f"[严重] 有 {invalid_date_count} 行账期日期无法识别（含 1970 年空值）。")

    stats = {
        "计算订单行数": len(kept_rows),
        "独立订单编号数": len(seen_order_ids),
        "跨月订单行数": len(cross_month_rows),
        "按规则剔除行数": len(dropped_rows),
        "订单处理费为0或为空": zero_fee_count,
        "燃油订单数量": fuel_order_count,
        "无燃油订单数量": no_fuel_order_count,
        "账期规则兜底行数": period_fallback_count,
        "物流类型分布": logistics_type_counter,
        "重算费用总计": decimal_to_float(round_money(total_recalc, digits, rmode)),
        "源文件费用总计": decimal_to_float(round_money(total_source, digits, rmode)),
        "总计差额": decimal_to_float(round_money(total_recalc - total_source, digits, rmode)),
    }

    return OutboundResult(
        headers=headers, rows=kept_rows, order_check_rows=order_check_rows,
        cross_month_rows=cross_month_rows, dropped_rows=dropped_rows, stats=stats,
    )


# ------------------------- 运行前检查（建议 十一） -------------------------

@dataclass
class PrecheckReport:
    items: List[Tuple[str, str, str]] = dc_field(default_factory=list)

    def ok(self, name: str, detail: str):
        self.items.append((name, str(detail), "通过"))

    def warn(self, name: str, detail: str):
        self.items.append((name, str(detail), "提示"))

    def block(self, name: str, detail: str):
        self.items.append((name, str(detail), "严重"))

    @property
    def blocked(self) -> bool:
        return any(level == "严重" for _n, _d, level in self.items)

    @property
    def has_warning(self) -> bool:
        return any(level == "提示" for _n, _d, level in self.items)

    def as_text(self) -> str:
        marks = {"通过": "√", "提示": "!", "严重": "×"}
        return "\n".join(f"  [{marks.get(lv, '-')}] {n}：{d}" for n, d, lv in self.items)

    def to_rows(self) -> List[Dict[str, str]]:
        return [{"检查项": n, "结果": d, "级别": lv} for n, d, lv in self.items]


def run_precheck(
    config: Dict[str, Any],
    merge_result: sm.MergeResult,
    year: int,
    month: int,
    periods: Sequence[fr.FuelPeriod],
    main_path: str,
    supplement_paths: Sequence[str],
    expected_order_count: Optional[int] = None,
) -> PrecheckReport:
    report = PrecheckReport()
    headers = merge_result.headers

    problems = cfgmod.validate_config(config)
    if problems:
        for p in problems:
            report.warn("客户配置", p)
    else:
        report.ok("客户配置", f"{config['customer_code']}（{config['customer_name']}）校验通过")

    report.ok("账单月份", f"{year}-{month:02d}")
    report.ok("主文件数量", f"1 个：{os.path.basename(main_path)}")
    report.ok("补充文件数量",
              f"{len(supplement_paths)} 个" +
              (f"：{'、'.join(os.path.basename(p) for p in supplement_paths)}" if supplement_paths else ""))

    for note in merge_result.structure_notes:
        report.warn("字段结构", note)

    critical = [config["fields"]["order_id"], config["fields"]["logistics_type"]]
    missing_critical = [f for f in critical if sm.resolve_field(headers, f) is None]
    if missing_critical:
        report.block("关键字段", f"缺少字段：{'、'.join(missing_critical)}")
    else:
        report.ok("关键字段", "订单编号 / 物流类型 均存在")

    period_fields = set(config["period_rules"].values()) | {config["period_default_field"]}
    missing_period = [f for f in sorted(period_fields) if sm.resolve_field(headers, f) is None]
    if missing_period:
        report.warn("账期时间字段", f"以下字段不存在，对应物流类型将无法判断账期：{'、'.join(missing_period)}")
    else:
        report.ok("账期时间字段", "、".join(sorted(period_fields)))

    money_fields = list(config["total_fields"]) + list(config["fuel_base_fields"]) + list(config["platform_fee_fields"])
    absent_money = [f for f in money_fields if sm.resolve_field(headers, f) is None]
    if absent_money:
        report.warn("金额字段", f"配置中 {len(absent_money)} 个字段本表不存在（按 0 计）：{'、'.join(absent_money)}")
    else:
        report.ok("金额字段", "配置的费用字段全部存在")

    text_cells: List[str] = []
    for row in merge_result.rows:
        for name in money_fields:
            actual = sm.resolve_field(headers, name)
            if actual and looks_like_money_text(row.get(actual)):
                text_cells.append(
                    f"{name}=「{row.get(actual)}」({row.get(sm.COL_SOURCE_FILE)} 第{row.get(sm.COL_SOURCE_ROW)}行)")
                break
        if len(text_cells) >= 10:
            break
    if text_cells:
        report.block("金额列内容", f"金额列中出现无法解析的文本：{'；'.join(text_cells)}")
    else:
        report.ok("金额列内容", "未发现文本型金额")

    severe_dup = [d for d in merge_result.duplicate_records if str(d.get("判定", "")).startswith("★")]
    if severe_dup:
        report.block("重复业务主键", f"{len(severe_dup)} 组主键重复但内容不同，需人工确认（见「重复订单」表）")
    else:
        report.ok("重复订单数量", f"{len(merge_result.duplicate_records)} 条完全一致的重复记录已丢弃")
    report.ok("补充新增订单数量", f"{merge_result.added_from_supplement} 条")
    report.ok("补充覆盖记录数量", f"{len(merge_result.override_records)} 条")

    type_counter: Dict[str, int] = {}
    cross_month = 0
    order_dates: List[date] = []
    invalid_dates = 0
    for row in merge_result.rows:
        ltype = str(sm.get_value(row, headers, config["fields"]["logistics_type"]) or "").strip() or "（空）"
        type_counter[ltype] = type_counter.get(ltype, 0) + 1
        pfield, _ = cfgmod.period_field_for(config, ltype if ltype != "（空）" else "")
        d = coerce_excel_date(sm.get_value(row, headers, pfield))
        if d is None or d.year < 2000:
            invalid_dates += 1
            continue
        order_dates.append(d)
        if not (d.year == year and d.month == month):
            cross_month += 1

    report.ok("物流类型分布", "、".join(f"{k}:{v}" for k, v in sorted(type_counter.items())) or "（无数据）")
    report.ok("跨月订单数量", f"{cross_month} 条（剔除并留档到「跨月订单」表）")
    if invalid_dates:
        report.block("账期日期", f"{invalid_dates} 行的账期时间无法识别或为 1970 年，需先在源文件中修正")
    else:
        report.ok("账期日期", "全部可识别")

    fuel_mode = config["fuel"]["mode"]
    if fuel_mode == "none":
        report.ok("燃油模式", "该客户不计燃油（fuel.mode=none），无需填写费率")
    elif fuel_mode == "source":
        report.ok("燃油模式", f"直接使用源文件字段【{config['fields']['source_fuel_fee']}】的燃油金额")
    else:
        billing_dates = [d for d in order_dates if d.year == year and d.month == month]
        rate_problems = fr.validate_periods(periods, required_dates=billing_dates)
        if rate_problems:
            for p in rate_problems:
                report.block("燃油费率表", p)
        else:
            report.ok("燃油费率表", f"{len(periods)} 个区间，无重叠无空档，已覆盖全部应计订单日期")

    if expected_order_count is not None:
        actual = len(merge_result.rows) - cross_month - invalid_dates
        if actual == expected_order_count:
            report.ok("系统订单件数核对", f"一致（{actual} 条）")
        else:
            report.warn("系统订单件数核对",
                        f"系统 {expected_order_count} 条 / 合并后本账期约 {actual} 条，"
                        f"相差 {actual - expected_order_count} 条")

    return report


# ------------------------- 其余 sheet 的汇总（沿用旧版规则） -------------------------

def process_warehouse_rent(wb, formula_function: str, log: LogFunc = _noop_log) -> Optional[FeeSummaryItem]:
    if SHEET_WAREHOUSE_RENT not in wb.sheetnames:
        log(f"[跳过] 未找到 sheet：{SHEET_WAREHOUSE_RENT}")
        return None

    ws = wb[SHEET_WAREHOUSE_RENT]
    amount_col = find_header_col(ws, RENT_AMOUNT_HEADER)
    if amount_col:
        log(f"  [{SHEET_WAREHOUSE_RENT}] 定位到「{RENT_AMOUNT_HEADER}」列：{get_column_letter(amount_col)}")
    else:
        amount_col = column_index_from_string(RENT_AMOUNT_FALLBACK_LETTER)
        log(f"  [{SHEET_WAREHOUSE_RENT}] 未找到表头「{RENT_AMOUNT_HEADER}」，回退固定列 {RENT_AMOUNT_FALLBACK_LETTER}，请核对表头。")
    total_cell = write_total_formula(ws, amount_col=amount_col, total_label="实际费用合计",
                                     formula_function=formula_function, key_cols=[amount_col])
    log(f"[完成] {SHEET_WAREHOUSE_RENT}!{total_cell} 已写入实际费用汇总。")
    return FeeSummaryItem("仓租费", SHEET_WAREHOUSE_RENT, total_cell, RENT_AMOUNT_HEADER)


def process_inbound_detail(wb, formula_function: str, log: LogFunc = _noop_log) -> Optional[FeeSummaryItem]:
    if SHEET_INBOUND_DETAIL not in wb.sheetnames:
        log(f"[跳过] 未找到 sheet：{SHEET_INBOUND_DETAIL}")
        return None

    ws = wb[SHEET_INBOUND_DETAIL]
    amount_col = find_header_col(ws, INBOUND_AMOUNT_HEADER)
    if amount_col:
        log(f"  [{SHEET_INBOUND_DETAIL}] 定位到「{INBOUND_AMOUNT_HEADER}」列：{get_column_letter(amount_col)}")
    else:
        amount_col = column_index_from_string(INBOUND_AMOUNT_FALLBACK_LETTER)
        log(f"  [{SHEET_INBOUND_DETAIL}] 未找到表头「{INBOUND_AMOUNT_HEADER}」，回退固定列 {INBOUND_AMOUNT_FALLBACK_LETTER}。")
    total_cell = write_total_formula(ws, amount_col=amount_col, total_label="费用合计",
                                     formula_function=formula_function, key_cols=[amount_col])
    log(f"[完成] {SHEET_INBOUND_DETAIL}!{total_cell} 已写入费用汇总。")
    return FeeSummaryItem("入库订单费", SHEET_INBOUND_DETAIL, total_cell, INBOUND_AMOUNT_HEADER)


ColumnChoiceCallback = Callable[[str, List[Tuple[int, str]]], Optional[Tuple[int, str]]]


def choose_candidate_column(ws, candidates, mode, column_choice_callback=None, log=_noop_log):
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]
    if mode == "skip":
        return None
    if mode == "ask" and column_choice_callback is not None:
        return column_choice_callback(ws.title, candidates)
    for keyword in UNKNOWN_AMOUNT_KEYWORDS:
        for col, header in candidates:
            if keyword == header or keyword in header:
                return col, header
    return candidates[0]


def process_other_sheets(
    wb, formula_function: str, already_processed_sheets: Iterable[str], other_sheet_mode: str,
    main_summary_sheet: str, column_choice_callback=None, log: LogFunc = _noop_log,
) -> List[FeeSummaryItem]:
    processed = set(already_processed_sheets)
    results: List[FeeSummaryItem] = []
    other_summary_rows: List[List[str]] = []

    for sheet_name in list(wb.sheetnames):
        if sheet_name in processed or sheet_name in REPORT_SHEETS:
            continue
        if is_summary_or_archive_sheet(sheet_name, main_summary_sheet):
            continue

        if sheet_name in SHEETS_ALREADY_COUNTED:
            covering_sheet, reason = SHEETS_ALREADY_COUNTED[sheet_name]
            if covering_sheet in processed:
                log(f"[跳过] {sheet_name}：{reason}")
                other_summary_rows.append([sheet_name, "", "", "", "跳过", reason])
                continue
            log(f"[注意] {sheet_name}：通常由【{covering_sheet}】计入，但本文件未处理该表，仍照常汇总以免漏算。")

        ws = wb[sheet_name]

        if sheet_name in FIXED_OTHER_SHEET_RULES:
            rule = FIXED_OTHER_SHEET_RULES[sheet_name]
            target_header = rule["target_header"]
            amount_col = find_header_col(ws, target_header)
            if not amount_col:
                msg = f"固定规则要求汇总「{target_header}」，但未找到该列，已跳过。"
                log(f"[跳过] {sheet_name}：{msg}")
                other_summary_rows.append([sheet_name, "", "", "", "跳过", msg])
                continue
            total_cell = write_total_formula(ws, amount_col=amount_col, total_label=f"{target_header}合计",
                                             formula_function=formula_function, key_cols=[amount_col])
            display_name = rule["display_name"]
            results.append(FeeSummaryItem(display_name, sheet_name, total_cell, target_header))
            other_summary_rows.append([sheet_name, display_name, target_header, total_cell, "已汇总", "固定规则"])
            log(f"[完成] {sheet_name}!{total_cell} 已按固定规则汇总「{target_header}」。")
            continue

        candidates = find_header_col_by_keywords(ws, UNKNOWN_AMOUNT_KEYWORDS)
        chosen = choose_candidate_column(ws, candidates, other_sheet_mode,
                                         column_choice_callback=column_choice_callback, log=log)
        if not chosen:
            msg = "未知 sheet 未找到可汇总费用列，或已选择跳过。"
            log(f"[跳过] {sheet_name}：{msg}")
            other_summary_rows.append([sheet_name, "", "", "", "跳过", msg])
            continue

        amount_col, header = chosen
        total_cell = write_total_formula(ws, amount_col=amount_col, total_label=f"{header}合计",
                                         formula_function=formula_function, key_cols=[amount_col])
        results.append(FeeSummaryItem(sheet_name, sheet_name, total_cell, header, "未知 sheet 通用识别"))
        other_summary_rows.append([sheet_name, sheet_name, header, total_cell, "已汇总", "通用识别"])
        log(f"[完成] {sheet_name}!{total_cell} 已按通用规则汇总「{header}」。")

    write_other_fee_summary_sheet(wb, other_summary_rows)
    return results


def write_other_fee_summary_sheet(wb, rows: List[List[str]]) -> None:
    if SHEET_OTHER_SUMMARY in wb.sheetnames:
        del wb[SHEET_OTHER_SUMMARY]
    ws = wb.create_sheet(SHEET_OTHER_SUMMARY)
    headers = ["来源sheet", "汇总显示名称", "汇总列", "合计单元格", "处理状态", "备注"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row_idx, col_idx).value = value
    for col_idx, width in enumerate([18, 18, 18, 14, 12, 36], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ------------------------- 汇总 sheet -------------------------

def sort_main_summary_items(items: Sequence[FeeSummaryItem]) -> List[FeeSummaryItem]:
    order = {name: idx for idx, name in enumerate(MAIN_SUMMARY_DISPLAY_ORDER)}
    seen, unique_items = set(), []
    for item in items:
        key = (item.display_name, item.sheet_name, item.total_cell)
        if key in seen:
            continue
        seen.add(key)
        unique_items.append(item)
    return sorted(unique_items, key=lambda i: (order.get(i.display_name, 999), i.display_name))


def _find_label_row(ws: Worksheet, label: str, limit_row: Optional[int] = None) -> Optional[int]:
    """在【汇总】sheet 的 A 列里按标签文字找行号（只找第一段，不动第二段物流账户部分）。"""
    target = norm_text(label)
    last = limit_row or ws.max_row
    for row in range(1, last + 1):
        if norm_text(ws.cell(row, 1).value) == target:
            return row
    return None


def _first_section_end(ws: Worksheet) -> int:
    """模板第 20 行左右是空行，之后是"物流账户"第二段。返回第一段的结束行号。"""
    row = _find_label_row(ws, SUMMARY_LABEL_CLOSING_BALANCE, ws.max_row)
    return row if row else ws.max_row


def find_or_create_main_summary_sheet(wb, summary_sheet_name: str):
    if summary_sheet_name in wb.sheetnames:
        return wb[summary_sheet_name]
    for ws in wb.worksheets:
        if norm_text(ws["A1"].value) == "类别" and norm_text(ws["B1"].value) == "金额":
            return ws
    ws = wb.create_sheet(summary_sheet_name, 0)
    ws["A1"] = "类别"
    ws["B1"] = "金额"
    for row, label in enumerate(
        ["客户编号", "客户昵称", "公司名称", "账单币种", "账单起止日期",
         SUMMARY_LABEL_OPENING_BALANCE, SUMMARY_LABEL_TOTAL_EXPENSE,
         *SUMMARY_EXPENSE_LABELS, SUMMARY_LABEL_RECHARGE, SUMMARY_LABEL_REFUND,
         SUMMARY_LABEL_CLOSING_BALANCE],
        start=2,
    ):
        ws.cell(row, 1).value = label
    return ws


def write_main_summary_sheet(wb, items, summary_sheet_name=DEFAULT_MAIN_SUMMARY_SHEET, log=_noop_log) -> None:
    """
    把各 sheet 的合计写回【汇总】模板。

    与旧版的关键差异：**按 A 列标签文字定位行**，而不是从第 7 行起顺序覆盖。
    旧写法会把模板里的「账单初账户余额」「减 本期账单总支出」等行冲掉。
    模板里没有对应标签的项目，会追加到第一段末尾并在日志中提示。
    """
    ws = find_or_create_main_summary_sheet(wb, summary_sheet_name)
    section_end = _first_section_end(ws)

    by_label: Dict[str, FeeSummaryItem] = {}
    for item in sort_main_summary_items(items):
        by_label.setdefault(item.display_name, item)

    def link_formula(item: FeeSummaryItem) -> str:
        cell_ref = re.sub(r"([A-Z]+)([0-9]+)", r"\1$\2", item.total_cell)
        return f"={quote_sheet_name(item.sheet_name)}!${cell_ref}"

    written: List[str] = []
    expense_rows: List[int] = []

    # ---- 1) 八项费用 + 充值 + 退款，按标签写入 ----
    for label in MAIN_SUMMARY_DISPLAY_ORDER:
        row = _find_label_row(ws, label, section_end)
        if row is None:
            continue
        item = by_label.get(label)
        cell = ws.cell(row, 2)
        cell.value = link_formula(item) if item else 0
        cell.number_format = "#,##0.00;[Red]-#,##0.00"
        if label in SUMMARY_EXPENSE_LABELS:
            expense_rows.append(row)
        if item:
            written.append(label)

    # ---- 2) 模板里没有的项目（例如旧版账单的「索赔抵扣」）追加到第一段末尾 ----
    extra = [lbl for lbl in by_label if lbl not in MAIN_SUMMARY_DISPLAY_ORDER]
    if extra:
        insert_at = section_end
        for offset, label in enumerate(sorted(extra)):
            ws.insert_rows(insert_at + offset)
            ws.cell(insert_at + offset, 1).value = label
            ws.cell(insert_at + offset, 2).value = link_formula(by_label[label])
            ws.cell(insert_at + offset, 2).number_format = "#,##0.00;[Red]-#,##0.00"
            expense_rows.append(insert_at + offset)
            written.append(label)
        section_end += len(extra)
        log(f"[提示] 【汇总】模板里没有以下标签，已追加到第一段末尾：{'、'.join(sorted(extra))}")

    # ---- 3) 合计行：本期总支出 / 期末余额 ----
    total_row = _find_label_row(ws, SUMMARY_LABEL_TOTAL_EXPENSE, section_end)
    if total_row and expense_rows:
        refs = "+".join(f"B{r}" for r in sorted(expense_rows))
        cell = ws.cell(total_row, 2)
        cell.value = f"={refs}"
        cell.number_format = "#,##0.00;[Red]-#,##0.00"
        cell.font = Font(bold=True)

    open_row = _find_label_row(ws, SUMMARY_LABEL_OPENING_BALANCE, section_end)
    recharge_row = _find_label_row(ws, SUMMARY_LABEL_RECHARGE, section_end)
    refund_row = _find_label_row(ws, SUMMARY_LABEL_REFUND, section_end)
    closing_row = _find_label_row(ws, SUMMARY_LABEL_CLOSING_BALANCE, section_end)

    if closing_row:
        parts = [f"B{r}" for r in (open_row, total_row, recharge_row, refund_row) if r]
        cell = ws.cell(closing_row, 2)
        cell.value = "=" + "+".join(parts) if parts else 0
        cell.number_format = "#,##0.00;[Red]-#,##0.00"
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFFF00")

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 20)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 16)

    missing = [lbl for lbl in by_label if lbl not in written]
    if missing:
        log(f"[注意] 以下汇总项没能写入【汇总】：{'、'.join(missing)}")
    log(f"[完成] 【{ws.title}】已按模板标签更新 {len(written)} 项费用，"
        f"「{SUMMARY_LABEL_TOTAL_EXPENSE}」「{SUMMARY_LABEL_CLOSING_BALANCE}」已重算，其余行未改动。")


def configure_excel_recalculation(wb) -> None:
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass


def write_fuel_rate_sheet(wb, periods: Sequence[fr.FuelPeriod]) -> None:
    if SHEET_FUEL_RATE in wb.sheetnames:
        del wb[SHEET_FUEL_RATE]
    ws = wb.create_sheet(SHEET_FUEL_RATE)
    headers = ["序号", "承运商", "服务类型", "开始日期", "结束日期", "燃油费率", "时区"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for idx, p in enumerate(sorted(periods, key=lambda x: x.start), start=1):
        ws.cell(idx + 1, 1).value = idx
        ws.cell(idx + 1, 2).value = p.carrier
        ws.cell(idx + 1, 3).value = p.service_group
        ws.cell(idx + 1, 4).value = p.start
        ws.cell(idx + 1, 5).value = p.end
        ws.cell(idx + 1, 6).value = float(p.rate)
        ws.cell(idx + 1, 7).value = p.timezone
        ws.cell(idx + 1, 4).number_format = "yyyy-mm-dd"
        ws.cell(idx + 1, 5).number_format = "yyyy-mm-dd"
        ws.cell(idx + 1, 6).number_format = "0.00%"
    for col, width in zip("ABCDEFG", [6, 12, 12, 14, 14, 12, 16]):
        ws.column_dimensions[col].width = width


# ------------------------- 其他 sheet 的月份清洗 -------------------------

def filter_dataframe_by_month(df: pd.DataFrame, date_col: str, year: int, month: int,
                              sheet_name: str, log: LogFunc = _noop_log):
    if date_col not in df.columns:
        log(f"[提示] 【{sheet_name}】中没有表头【{date_col}】，跳过按月筛选。")
        return df, df.iloc[0:0].copy()

    parsed = pd.to_datetime(df[date_col], errors="coerce")
    mask = (parsed.dt.year == year) & (parsed.dt.month == month)
    removed = df[~mask].copy()
    clean = df[mask].copy()
    if len(removed) > 0:
        removed.insert(0, "删除原因", f"非{year}-{month:02d}数据（按{date_col}判断）")
        log(f"[{sheet_name}] 非{year}-{month:02d}的数据共 {len(removed)} 条，已剔除并留档。")
    else:
        log(f"[{sheet_name}] 全部数据均属于{year}-{month:02d}。")
    return clean, removed


def dataframe_is_empty(df: pd.DataFrame) -> bool:
    if df is None or len(df.columns) == 0:
        return True
    return len(df.dropna(how="all")) == 0


def drop_header_only_sheets(all_sheets: Dict[str, pd.DataFrame], protected=None, log=_noop_log):
    protected_set = set(protected) if protected is not None else set(PROTECTED_SHEETS_FROM_REMOVAL)
    kept, removed = {}, []
    for name, df in all_sheets.items():
        if dataframe_is_empty(df):
            if name in protected_set:
                log(f"  [保留] 【{name}】仅有表头，但属于必需sheet，保留（合计记为0）。")
                kept[name] = df
            else:
                removed.append(name)
        else:
            kept[name] = df
    if removed:
        log(f"[空sheet清除] 已自动删除 {len(removed)} 个仅有表头的sheet：{'、'.join(removed)}")
    else:
        log("[空sheet清除] 未发现仅有表头的空sheet。")
    return kept, removed


# ------------------------- 输出 -------------------------

def _write_sheets_to_excel(output_sheets: Dict[str, pd.DataFrame], path: str) -> None:
    try:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for name, df in output_sheets.items():
                df.to_excel(writer, sheet_name=str(name)[:31], index=False)
    except PermissionError as exc:
        raise BillingProcessError(
            f"无法写入文件：{path}\n该文件可能正被 Excel 打开，请关闭后重试。") from exc


def _save_workbook(wb, output_path: str) -> None:
    try:
        wb.save(output_path)
    except PermissionError as exc:
        raise BillingProcessError(
            f"无法保存结果文件：{output_path}\n该文件可能正被 Excel 打开，请关闭后重新运行。") from exc
    except OSError as exc:
        raise BillingProcessError(f"保存结果文件失败：{output_path}\n详细信息：{exc}") from exc


def _df(rows: List[Dict[str, Any]], columns: Optional[Sequence[str]] = None) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=list(columns) if columns else ["（本次无记录）"])
    return pd.DataFrame(rows, columns=list(columns) if columns else None)


def read_summary_template_values(main_path: str,
                                 summary_sheet: str = DEFAULT_MAIN_SUMMARY_SHEET) -> Dict[str, Any]:
    """读出原账单【汇总】模板里各标签的原始金额，用于和明细重算值对账。"""
    values: Dict[str, Any] = {}
    try:
        wb = load_workbook(main_path, data_only=True)
    except Exception:
        return values
    try:
        if summary_sheet not in wb.sheetnames:
            return values
        ws = wb[summary_sheet]
        stop = ws.max_row
        for row in range(1, stop + 1):
            label = norm_text(ws.cell(row, 1).value)
            if not label:
                continue
            if label in values:      # 只取第一段（第二段是物流账户，标签会重复）
                continue
            values[label] = ws.cell(row, 2).value
    finally:
        wb.close()
    return values


def _dataframe_column_total(df: pd.DataFrame, header: str) -> Optional[Decimal]:
    """把某一列的数值求和（跳过"合计"之类的非数值行）。"""
    if df is None or header not in df.columns:
        return None
    total = ZERO
    for value in df[header]:
        if is_blank_value(value):
            continue
        try:
            total += to_decimal(value, strict=True)
        except MoneyParseError:
            continue
    return total


def append_reconciliation(wb, all_sheets: Dict[str, pd.DataFrame],
                          summary_items: Sequence[FeeSummaryItem],
                          original_summary: Dict[str, Any],
                          digits: int = 2, log: LogFunc = _noop_log) -> List[Dict[str, Any]]:
    """
    在【核对汇总】里追加"账单原值 vs 明细重算值"的对账表。

    这是发现"漏单"的关键：若账单【汇总】里的出库订单费比【出库订单】明细合计更多，
    说明有订单只计了钱、明细表里却没有对应行，需要用补充文件补齐。
    """
    rows: List[Dict[str, Any]] = []
    for item in summary_items:
        recalc = _dataframe_column_total(all_sheets.get(item.sheet_name), item.amount_header)
        original = original_summary.get(norm_text(item.display_name))
        original_dec = None
        if not is_blank_value(original):
            try:
                original_dec = round_money(to_decimal(original, strict=True), digits)
            except MoneyParseError:
                original_dec = None
        recalc_dec = round_money(recalc, digits) if recalc is not None else None
        diff = (recalc_dec - original_dec) if (recalc_dec is not None and original_dec is not None) else None

        rows.append({
            "项目": item.display_name,
            "来源sheet": item.sheet_name,
            "汇总列": item.amount_header,
            "账单原值": decimal_to_float(original_dec) if original_dec is not None else "",
            "明细重算值": decimal_to_float(recalc_dec) if recalc_dec is not None else "",
            "差额": decimal_to_float(diff) if diff is not None else "",
            "说明": ("一致" if diff == ZERO else
                     "★账单金额与明细表合计不符，可能存在漏单或明细缺行，请用补充文件补齐"
                     if diff is not None else "无法比对"),
        })

    if SHEET_CHECK_SUMMARY not in wb.sheetnames:
        return rows

    ws = wb[SHEET_CHECK_SUMMARY]
    start = ws.max_row + 2
    ws.cell(start, 1).value = "【账单原值 vs 明细重算值 对账】"
    ws.cell(start, 1).font = Font(bold=True)
    headers = ["项目", "来源sheet", "汇总列", "账单原值", "明细重算值", "差额", "说明"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start + 1, col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for offset, data in enumerate(rows, start=start + 2):
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(offset, col)
            cell.value = data[header]
            if header in ("账单原值", "明细重算值", "差额"):
                cell.number_format = "#,##0.00;[Red]-#,##0.00"
            if header == "说明" and str(data[header]).startswith("★"):
                cell.font = Font(color="CC0000", bold=True)
    for col, width in enumerate([18, 14, 14, 16, 16, 14, 52], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    flagged = [r for r in rows if str(r["说明"]).startswith("★")]
    for r in flagged:
        log(f"[注意] 【{r['项目']}】账单原值 {r['账单原值']}，明细表合计 {r['明细重算值']}，"
            f"差额 {r['差额']}。详见「核对汇总」。")
    return rows


def finalize_outbound_sheet(wb, config, formula_function: str, log=_noop_log) -> Optional[FeeSummaryItem]:
    """出库订单每单金额已由 Python 用 Decimal 算好并写成数值，这里只补合计行与数字格式。"""
    if SHEET_OUTBOUND_ORDER not in wb.sheetnames:
        log(f"[跳过] 未找到 sheet：{SHEET_OUTBOUND_ORDER}")
        return None

    ws = wb[SHEET_OUTBOUND_ORDER]
    total_header = config["fields"]["total"]
    total_col = find_header_col(ws, total_header)
    if not total_col:
        raise BillingProcessError(f"【{SHEET_OUTBOUND_ORDER}】中找不到表头「{total_header}」，无法生成合计。")

    for header in (config["fields"]["fuel_fee"], config["fields"]["logistics_fee"], total_header):
        col = find_header_col(ws, header)
        if col:
            for row in range(2, ws.max_row + 1):
                ws.cell(row, col).number_format = '#,##0.00;[Red]-#,##0.00'

    total_cell = write_total_formula(ws, amount_col=total_col, total_label="费用总计合计",
                                     formula_function=formula_function, key_cols=[total_col])
    log(f"[完成] {SHEET_OUTBOUND_ORDER}!{total_cell} 已写入费用总计汇总。")
    return FeeSummaryItem("出库订单费", SHEET_OUTBOUND_ORDER, total_cell, total_header)


# ------------------------- 主流程 -------------------------

OrderMismatchCallback = Callable[[int, int], bool]
WeeklyRateCallback = Callable[[List[Tuple[date, date]]], Optional[Dict[date, Decimal]]]
PrecheckCallback = Callable[[PrecheckReport], bool]


def month_bounds(year: int, month: int) -> Tuple[date, date]:
    return date(year, month, 1), date(year, month, calendar.monthrange(year, month)[1])


def collect_outbound_dates(merge_result: sm.MergeResult, config: Dict[str, Any]) -> List[date]:
    dates: List[date] = []
    for row in merge_result.rows:
        ltype = str(sm.get_value(row, merge_result.headers, config["fields"]["logistics_type"]) or "").strip()
        pfield, _ = cfgmod.period_field_for(config, ltype)
        d = coerce_excel_date(sm.get_value(row, merge_result.headers, pfield))
        if d and d.year >= 2000:
            dates.append(d)
    return dates


def build_week_ranges_for(dates: Sequence[date], year: int, month: int,
                          week_start_weekday: int = 0) -> List[Tuple[date, date]]:
    """算出覆盖本账期所有订单日期的费率周区间（不做任何 ±1 天平移）。"""
    start, end = month_bounds(year, month)
    in_month = [d for d in dates if d.year == year and d.month == month] or [start, end]
    return fr.carrier_week_ranges(min(min(in_month), start), max(max(in_month), end), week_start_weekday)


def week_ranges_for_config(merge_result, config, year: int, month: int) -> List[Tuple[date, date]]:
    return build_week_ranges_for(
        collect_outbound_dates(merge_result, config), year, month,
        week_start_weekday=config["fuel"].get("week_start_weekday", 0),
    )


def infer_rates_from_file(main_path: str, customer_code: str, year: int, month: int):
    """
    从源账单反推每个费率周实际使用的燃油费率（诊断用，不参与计算）。
    返回 [(start, end, 推断费率, 样本数, 说明), ...]
    """
    config = cfgmod.get_customer_config(customer_code)
    if config["fuel"]["mode"] == "none":
        return []

    main_sheet = sm.read_sheet(main_path, config["sheets"]["outbound"], role="主文件")
    headers = main_sheet.headers
    base_fields = [f for f in config["fuel_base_fields"] if sm.resolve_field(headers, f)]
    fuel_field = sm.resolve_field(headers, config["fields"]["source_fuel_fee"])
    if not base_fields or not fuel_field:
        return []

    rate_field = config["fuel"].get("rate_date_field", "创建时间")
    if rate_field in ("period", "", None):
        rate_field = config["period_default_field"]

    samples = []
    dates = []
    for row in main_sheet.rows:
        d = coerce_excel_date(sm.get_value(row, headers, rate_field))
        if d is None or d.year < 2000:
            continue
        dates.append(d)
        base = ZERO
        for f in base_fields:
            base += to_decimal(row.get(sm.resolve_field(headers, f)), strict=False)
        fuel = to_decimal(row.get(fuel_field), strict=False)
        samples.append((d, base, fuel))

    week_ranges = build_week_ranges_for(
        dates, year, month, week_start_weekday=config["fuel"].get("week_start_weekday", 0))
    return fr.infer_rates_from_samples(samples, week_ranges)


def preview_week_ranges(main_path: str, supplement_paths: Sequence[str],
                        customer_code: str, year: int, month: int) -> List[Tuple[date, date]]:
    """供 GUI 在弹出「按周填写费率」对话框前，先算出需要填写哪些周。"""
    config = cfgmod.get_customer_config(customer_code)
    main_sheet = sm.read_sheet(main_path, config["sheets"]["outbound"], role="主文件")
    supplements = [sm.read_sheet(p, config["sheets"]["outbound"], role="补充文件")
                   for p in supplement_paths]
    merge_result = sm.merge_sources(
        main_sheet, supplements,
        key_fields=config["merge"]["key"],
        supplement_priority=config["merge"]["supplement_priority"],
        allow_extra_fields=config["merge"]["allow_extra_fields"],
        strict_structure=False,
    )
    return week_ranges_for_config(merge_result, config, year, month)


def run_pipeline(
    main_path: str,
    year: int,
    month: int,
    output_path: str,
    customer_code: str = "DEFAULT",
    supplement_paths: Optional[Sequence[str]] = None,
    uniform_fuel_rate: Optional[Decimal] = None,
    fuel_periods: Optional[Sequence[fr.FuelPeriod]] = None,
    fuel_rate_csv: Optional[str] = None,
    weekly_rate_callback: Optional[WeeklyRateCallback] = None,
    expected_order_count: Optional[int] = None,
    formula_function: str = "SUM",
    other_sheet_mode: str = "auto",
    remove_empty_sheets: bool = True,
    column_choice_callback: Optional[ColumnChoiceCallback] = None,
    order_count_mismatch_callback: Optional[OrderMismatchCallback] = None,
    precheck_callback: Optional[PrecheckCallback] = None,
    log: LogFunc = _noop_log,
) -> dict:
    """完整流程：读取主文件 + 补充文件 -> 结构校验 -> 合并 -> 运行前检查 -> 逐单核算 -> 输出。"""
    supplement_paths = list(supplement_paths or [])
    config = cfgmod.get_customer_config(customer_code)
    collector = IssueCollector(config["customer_code"])

    log(f"客户编码：{config['customer_code']}（{config['customer_name']}），账期 {year}-{month:02d}")
    log(f"燃油模式：{config['fuel']['mode']}")

    # ---- 1/2. 读取并合并 ----
    outbound_sheet_name = config["sheets"]["outbound"]
    log(f"正在读取主文件：{os.path.basename(main_path)}")
    try:
        main_sheet = sm.read_sheet(main_path, outbound_sheet_name, role="主文件")
    except sm.StructureError as exc:
        raise BillingProcessError(str(exc)) from exc
    log(f"  主文件【{outbound_sheet_name}】：{len(main_sheet.headers)} 列 / {len(main_sheet.rows)} 行")

    supplements: List[sm.SourceSheet] = []
    for path in supplement_paths:
        log(f"正在读取补充文件：{os.path.basename(path)}")
        try:
            s = sm.read_sheet(path, outbound_sheet_name, role="补充文件")
        except sm.StructureError as exc:
            raise BillingProcessError(str(exc)) from exc
        log(f"  补充文件【{s.sheet_name}】：{len(s.headers)} 列 / {len(s.rows)} 行")
        supplements.append(s)

    try:
        merge_result = sm.merge_sources(
            main_sheet, supplements,
            key_fields=config["merge"]["key"],
            supplement_priority=config["merge"]["supplement_priority"],
            allow_extra_fields=config["merge"]["allow_extra_fields"],
            strict_structure=config["merge"]["strict_structure"],
            watch_fields=config["total_fields"],
        )
    except sm.StructureError as exc:
        raise BillingProcessError(str(exc)) from exc

    log(f"合并完成：共 {len(merge_result.rows)} 行"
        f"（补充新增 {merge_result.added_from_supplement} 条，"
        f"覆盖 {len(merge_result.override_records)} 条，"
        f"重复丢弃 {len(merge_result.duplicate_records)} 条）")
    for note in merge_result.structure_notes:
        log(f"  [结构提示] {note}")

    # ---- 燃油费率区间 ----
    periods: List[fr.FuelPeriod] = []
    if config["fuel"]["mode"] == "schedule":
        if fuel_periods:
            periods = list(fuel_periods)
        elif fuel_rate_csv:
            periods = fr.load_periods_from_csv(fuel_rate_csv)
            log(f"已从费率表载入 {len(periods)} 个区间：{fuel_rate_csv}")
        else:
            week_ranges = week_ranges_for_config(merge_result, config, year, month)
            if weekly_rate_callback is not None:
                log(f"本账期覆盖 {len(week_ranges)} 个费率周，等待按周填写燃油费率...")
                weekly = weekly_rate_callback(week_ranges)
                if weekly is None:
                    raise PipelineCancelled("用户取消了按周燃油费率的填写。")
                missing = [s for s, _e in week_ranges if s not in weekly]
                if missing:
                    raise BillingProcessError(
                        "缺少以下周的燃油费率：" + "、".join(f"{s:%Y-%m-%d}" for s in missing))
                periods = fr.build_periods_from_weekly_rates(
                    [(s, e, weekly[s]) for s, e in week_ranges], carrier=config["fuel"]["carrier"])
            elif uniform_fuel_rate is not None:
                periods = fr.build_periods_from_weekly_rates(
                    [(s, e, uniform_fuel_rate) for s, e in week_ranges], carrier=config["fuel"]["carrier"])
            else:
                raise BillingProcessError(
                    "该客户燃油模式为「按费率表计算」，请填写燃油费率、按周填写，或导入费率表 CSV。")
        log(f"燃油费率区间：{fr.describe_periods(periods)}")

    # ---- 运行前检查 ----
    log("正在执行运行前检查...")
    report = run_precheck(config, merge_result, year, month, periods,
                          main_path, supplement_paths, expected_order_count)
    log(report.as_text())

    if report.blocked:
        details = "\n".join(f"  - {n}：{d}" for n, d, lv in report.items if lv == "严重")
        raise PrecheckBlocked("运行前检查发现严重异常，已停止计算：\n" + details)

    if precheck_callback is not None and not precheck_callback(report):
        raise PipelineCancelled("用户在运行前检查环节选择了中止。")

    # ---- 逐单核算 ----
    log("正在逐单核算出库订单...")
    result = calculate_outbound(merge_result, config, year, month, periods, collector, log=log)

    blocking = collector.blocking()
    if blocking:
        preview = "\n".join(
            f"  - [{i.issue_type}] 订单 {i.order_id or '(空)'}（{i.source_file} 第{i.source_row}行）：{i.reason}"
            for i in blocking[:15])
        more = f"\n  ... 共 {len(blocking)} 条" if len(blocking) > 15 else ""
        raise BillingProcessError("存在必须先处理的严重数据异常，已停止计算：\n" + preview + more)

    log(f"核算完成：计入 {result.stats['计算订单行数']} 行，"
        f"跨月剔除 {result.stats['跨月订单行数']} 行，"
        f"规则剔除 {result.stats['按规则剔除行数']} 行。")
    log(f"重算费用总计 {result.stats['重算费用总计']}，"
        f"源文件费用总计 {result.stats['源文件费用总计']}，"
        f"差额 {result.stats['总计差额']}。")

    # ---- 件数核对 ----
    actual_order_count = result.stats["计算订单行数"]
    if expected_order_count is not None and expected_order_count != actual_order_count:
        msg = (f"核对不一致：系统订单件数 {expected_order_count}，"
               f"本账期计入 {actual_order_count} 行，相差 {actual_order_count - expected_order_count}。")
        log(msg)
        cont = order_count_mismatch_callback(actual_order_count, expected_order_count) \
            if order_count_mismatch_callback else False
        if not cont:
            raise OrderCountMismatchError(msg)
        log("用户确认继续。")

    # ---- 组装输出 ----
    log("正在生成结果文件...")
    original_summary = read_summary_template_values(main_path)
    all_sheets = pd.read_excel(main_path, sheet_name=None)
    removed_sheets: List[str] = []
    if remove_empty_sheets:
        all_sheets, removed_sheets = drop_header_only_sheets(all_sheets, log=log)

    if SHEET_WAREHOUSE_RENT in all_sheets:
        rent_clean, rent_removed = filter_dataframe_by_month(
            all_sheets[SHEET_WAREHOUSE_RENT], RENT_DATE_COL, year, month, SHEET_WAREHOUSE_RENT, log=log)
        all_sheets[SHEET_WAREHOUSE_RENT] = rent_clean
        if len(rent_removed) > 0:
            all_sheets[f"{SHEET_WAREHOUSE_RENT}_月份外已删除"] = rent_removed

    all_sheets[SHEET_OUTBOUND_ORDER] = _df(result.rows, result.headers)

    stats = result.stats
    check_summary_rows = [
        {"项目": "客户编码", "值": config["customer_code"]},
        {"项目": "账单月份", "值": f"{year}-{month:02d}"},
        {"项目": "主文件", "值": os.path.basename(main_path)},
        {"项目": "补充文件", "值": "、".join(os.path.basename(p) for p in supplement_paths) or "（无）"},
        {"项目": "燃油模式", "值": config["fuel"]["mode"]},
        {"项目": "燃油费率区间", "值": fr.describe_periods(periods) if periods else "（不适用）"},
        {"项目": "合并后总行数", "值": len(merge_result.rows)},
        {"项目": "补充新增订单", "值": merge_result.added_from_supplement},
        {"项目": "补充覆盖订单", "值": len(merge_result.override_records)},
        {"项目": "重复丢弃行数", "值": len(merge_result.duplicate_records)},
    ] + [
        {"项目": k, "值": ("、".join(f"{a}:{b}" for a, b in v.items()) if isinstance(v, dict) else v)}
        for k, v in stats.items()
    ]

    all_sheets[SHEET_CHECK_SUMMARY] = _df(check_summary_rows)
    all_sheets[SHEET_ORDER_CHECK] = _df(result.order_check_rows)
    all_sheets[SHEET_MISSING_ORDERS] = _df([r for r in result.order_check_rows if r.get("数据来源") == "补充文件"])
    all_sheets[SHEET_DUPLICATE_ORDERS] = _df(merge_result.duplicate_records)
    all_sheets[SHEET_CROSS_MONTH] = _df(result.cross_month_rows)
    all_sheets[SHEET_OVERRIDE_LOG] = _df(merge_result.override_records)
    all_sheets[SHEET_RATE_ISSUES] = _df(collector.by_category(SHEET_RATE_ISSUES))
    all_sheets[SHEET_FIELD_ISSUES] = _df(collector.by_category(SHEET_FIELD_ISSUES))
    all_sheets[SHEET_UNEXPLAINED] = _df(collector.by_category(SHEET_UNEXPLAINED))
    all_sheets[SHEET_MANUAL_ADJUST] = _df(result.dropped_rows)
    all_sheets[SHEET_DATA_QUALITY] = _df(report.to_rows())

    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)
    try:
        _write_sheets_to_excel(all_sheets, tmp_path)
        wb = load_workbook(tmp_path)

        summary_items: List[FeeSummaryItem] = []
        processed_sheets = set(REPORT_SHEETS)

        item = process_warehouse_rent(wb, formula_function, log=log)
        if item:
            summary_items.append(item)
            processed_sheets.add(item.sheet_name)

        item = process_inbound_detail(wb, formula_function, log=log)
        if item:
            summary_items.append(item)
            processed_sheets.add(item.sheet_name)

        item = finalize_outbound_sheet(wb, config, formula_function, log=log)
        if item:
            summary_items.append(item)
            processed_sheets.add(item.sheet_name)

        summary_items.extend(process_other_sheets(
            wb, formula_function=formula_function, already_processed_sheets=processed_sheets,
            other_sheet_mode=other_sheet_mode, main_summary_sheet=DEFAULT_MAIN_SUMMARY_SHEET,
            column_choice_callback=column_choice_callback, log=log))

        if periods:
            write_fuel_rate_sheet(wb, periods)

        write_main_summary_sheet(wb, summary_items, DEFAULT_MAIN_SUMMARY_SHEET, log=log)
        reconciliation = append_reconciliation(
            wb, all_sheets, summary_items, original_summary,
            digits=config["rounding"]["digits"], log=log)
        configure_excel_recalculation(wb)
        _save_workbook(wb, output_path)
        log(f"处理完成，结果已保存至：{output_path}")
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    return {
        "stage": "full",
        "customer_code": config["customer_code"],
        "output_path": output_path,
        "actual_order_count": actual_order_count,
        "expected_order_count": expected_order_count,
        "removed_empty_sheets": removed_sheets,
        "summary_items": summary_items,
        "stats": stats,
        "fuel_periods": periods,
        "reconciliation": reconciliation,
        "precheck": report,
        "issue_counts": {
            SHEET_RATE_ISSUES: collector.count(SHEET_RATE_ISSUES),
            SHEET_FIELD_ISSUES: collector.count(SHEET_FIELD_ISSUES),
            SHEET_UNEXPLAINED: collector.count(SHEET_UNEXPLAINED),
        },
    }
